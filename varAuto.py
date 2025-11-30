import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np
import json, re, unicodedata
from pathlib import Path
from io import BytesIO
from bs4 import BeautifulSoup
import html
from difflib import SequenceMatcher
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import sys
import os

# Sicherstellen, dass stderr existiert
if sys.stderr is None:
    # Erstelle Log-Verzeichnis (falls nicht vorhanden)
    log_dir = Path(__file__).parent / "logs"
    log_dir.mkdir(exist_ok=True)
    
    # Leite stderr und stdout in Log-Datei
    log_file = log_dir / f"varauto_{pd.Timestamp.now().strftime('%Y%m%d')}.log"
    sys.stderr = open(log_file, 'a', encoding='utf-8')
    sys.stdout = open(log_file, 'a', encoding='utf-8')
    
    print(f"📝 Alle Logs umgeleitet nach: {log_file}")
    
if sys.stderr is None:
    sys.stderr = open(log_file, 'a', encoding='utf-8')
if sys.stdout is None:
    sys.stdout = open(log_file, 'a', encoding='utf-8')

# Danach erst faulthandler aktivieren
import faulthandler
faulthandler.enable(file=sys.stderr)
    
import shutil
import tempfile
import spacy
from spacy.util import load_model_from_path
import random
import threading
import tiktoken
import traceback
from functools import partial
from openai import OpenAI
from TitleSelection import TitleSelectionDialog
from UIvarAuto import KeywordSelectionDialog
from ModuleNotFoundSelectionDialog import ModuleEditDialog
from DublettenCheck import GroupEditDialog, DuplicateResolutionDialog
from collections import Counter
import fasttext
import langid
from typing import Dict, Set, List, Tuple, Any, Optional
from msal import ConfidentialClientApplication
import math
import psutil, os
from PySide6.QtCore import Qt, QEventLoop, QTimer, QThread, Signal, QObject
from PySide6.QtGui import QPixmap
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QDialog, QTextEdit, QProgressBar, QPlainTextEdit,
    QPushButton, QLabel, QGroupBox, QSlider, QCheckBox, QSplashScreen, QListWidget, QMessageBox, QInputDialog,
    QFileDialog, QFormLayout, QDialogButtonBox, QLineEdit, QComboBox, QTabWidget, QFrame
)
from dataclasses import dataclass
import requests



class UserCancelledError(KeyboardInterrupt):
    """Wird ausgelöst, wenn der Benutzer in der GUI auf 'Abbrechen' klickt."""
    pass

class ModuleNotFoundError(ValueError):
    """Wird geworfen, wenn eine Modulnummer in keiner Spalte vorhanden ist."""
    pass

class BadModuleIds(ValueError):
    """Module list is unusable – let user pick the correct ones."""
    pass

class QHLine(QFrame):
    def __init__(self):
        super().__init__()
        self.setFrameShape(QFrame.Shape.HLine)
        self.setFrameShadow(QFrame.Shadow.Sunken)

class PromptManager:
    def __init__(self, filename="prompts.json", initial_data=None):
        self.filename = filename
        self.prompts = {}
        
        if initial_data is not None:
            self.prompts = initial_data
            self.save()

        elif os.path.exists(filename):
            self.load()
        else:
            self.prompts = {
                "translate_title_to_german": {"prompt": "", "temperature": 0.5},
                "translate_to_german": {"prompt": "", "temperature": 0.5},
                "ai_voraussetzungen": {"prompt": "", "temperature": 0.5},
                "ai_zielgruppe": {"prompt": "", "temperature": 0.5},
                "ai_einleitung": {"prompt": "", "temperature": 0.5},
                "ai_inhalte": {"prompt": "", "temperature": 0.5},
                "ai_keywords": {"prompt": "", "temperature": 0.5},
                "ai_abschlussart_primary": {"prompt": "", "temperature": 0.5},
                "enforce_length_via_ai": {"prompt": "", "temperature": 0.5},
                "llm_best_two_kursnet": {"prompt": "", "temperature": 0.5}
            }
            self.save()

    def load(self):
        with open(self.filename, "r", encoding="utf-8") as f:
            self.prompts = json.load(f)

    def save(self):
        with open(self.filename, "w", encoding="utf-8") as f:
            json.dump(self.prompts, f, ensure_ascii=False, indent=4)

    def __getitem__(self, key):
        return self.prompts.get(key, "")

    def __setitem__(self, key, value):
        self.prompts[key] = value
        

# ---- globaler Logger ----
class QtLogEmitter(QObject):
    log_signal = Signal(str)

log_emitter = QtLogEmitter()


def log(*args, sep=" ", end="\n"):
    """
    Alternative zu print(): sendet Text als Signal an MasterUI.
    Threadsafe. Keine direkte GUI-Abhängigkeit.
    """
    text = sep.join(str(a) for a in args) + end
    log_emitter.log_signal.emit(text.rstrip("\n"))

        
        
        
class PromptEditorDialog(QDialog):
    def __init__(self, prompt_manager: PromptManager, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Prompt-Editor")
        self.setMinimumSize(900, 550)

        self.pm = prompt_manager

        layout = QHBoxLayout(self)

        # Links: Liste aller Prompts
        self.list = QListWidget()
        for key in self.pm.prompts.keys():
            self.list.addItem(key)
        layout.addWidget(self.list, 1)

        # Rechts: Editor + Buttons
        right = QVBoxLayout()
        self.editor = QTextEdit()
        right.addWidget(self.editor, 10)

        # Buttons
        btn_row = QHBoxLayout()
        self.btn_save = QPushButton("Speichern")
        self.btn_load = QPushButton("Neu Laden")
        self.btn_close = QPushButton("Schließen")
        btn_row.addWidget(self.btn_load)
        btn_row.addWidget(self.btn_save)
        btn_row.addWidget(self.btn_close)
        right.addLayout(btn_row)

        layout.addLayout(right, 3)
        
        # Temperaturbereich
        temp_area = QHBoxLayout()
        self.temp_label = QLabel("Temperatur: 0.50")
        self.temp_slider = QSlider(Qt.Horizontal)
        self.temp_slider.setRange(0, 100)
        self.temp_slider.valueChanged.connect(self._on_temp_change)
        temp_area.addWidget(self.temp_label)
        temp_area.addWidget(self.temp_slider)
        right.addLayout(temp_area)


        # Events
        self.list.currentItemChanged.connect(self._on_list_select)
        self.btn_save.clicked.connect(self._on_save)
        self.btn_load.clicked.connect(self._on_reload)
        self.btn_close.clicked.connect(self.close)

        # Startzustand
        self.list.setCurrentRow(0)
        self._on_list_select()
        
        
        
    def _on_temp_change(self, val):
        self.temp_label.setText(f"Temperatur: {val / 100:.2f}")
        

    def _on_list_select(self):
        key = self.list.currentItem().text()
        value = self.pm[key]
        if isinstance(value, dict):
            self.editor.setPlainText(value.get("prompt", ""))
            temp = value.get("temperature", 0.5)
            self.temp_slider.setValue(int(temp * 100))
            self.temp_label.setText(f"Temperatur: {temp:.2f}")
        else:
            self.editor.setPlainText(value)


    def _on_save(self):
        key = self.list.currentItem().text()
        if isinstance(self.pm[key], dict):
            self.pm[key]["prompt"] = self.editor.toPlainText()
            self.pm[key]["temperature"] = self.temp_slider.value() / 100
        else:
            self.pm[key] = self.editor.toPlainText()
        self.pm.save()
        QMessageBox.information(self, "Gespeichert", f"Prompt '{key}' gespeichert.")


    def _on_reload(self):
        self.pm.load()
        self._on_list_select()
        QMessageBox.information(self, "Neu geladen", "Prompts erneut geladen.")
        
        
# ============================================================================
# AUTOMATISCHE CONFIG-DIAGNOSE & REPARATUR
# ============================================================================
def diagnose_and_fix_config():
    """
    Führt eine vollständige Diagnose durch:
    1. Überprüft alle Excel-Dateien auf Spalten-Konsistenz
    2. Korrigiert leere Config-Einträge automatisch
    3. Zeigt klare Fehlermeldungen bei kritischen Problemen
    """

    print("🔍 Starte automatische Config-Diagnose...")
    
    issues_found = []
    fixes_applied = []
    
    # ============================================================
    # 1. PRÜFE MODULKOMBI SHEET
    # ============================================================
    filename = config_manager.get("modulkombi.filename")
    if filename and os.path.exists(filename):
        try:
            # Lade nur Header
            df_header = pd.read_excel(filename, nrows=0)
            actual_columns = list(df_header.columns)
            
            print(f"📊 Modulkombi '{filename}':")
            print(f"   Gefundene Spalten: {len(actual_columns)}")
            
            # Prüfe jeden Config-Wert
            expected_fields = {
                "Modulnummern": "Modulnummern kommagetrennt NEU\nNach Übernahme ",
                "Kursübersicht": "Kursübersicht der Einzelmodule (NEU)",
                "EinleitungNeu": "Einleitung NEU Sabine vgl. Spalte I\n\nGrün: in ESTHER übernommen",
                "TitelFeld": "Neu zertifizierter Titel der Maßnahme NEU",
                "TitelNeu": "Titeloptimierung Edtelligent/ Sabine\nvgl. Spalte C rosa\nVgl. Bericht 5.8\nGrün: in ESTHER übernommen",
                "Systematik_Feld": "Systematik in Kursnet",
                "Keyword_Feld": "Keyword-Optimierung Sabine/ Edtelligent 04_25\n\nVgl. Bericht 5.6\nGrün: in ESTHER übernommen",
                "Termine_Feld": "mögliche Zeiträume für ESTHER \n(rot = keine Kombination gefunden, \ngrün = fertig) ",
                "Inhalte_Feld": "Inhalt NEU",
                "DauerFeld": "Dauer in Tagen",
            }
            
            for field, default_value in expected_fields.items():
                config_value = config_manager.get(f"modulkombi.columns.{field}")
                
                # Problem 1: Leerer Wert
                if not config_value or config_value.strip() == "":
                    issues_found.append(f"Leerer Config-Wert: modulkombi.columns.{field}")
                    # AUTO-FIX: Setze auf Default
                    config_manager.set(f"modulkombi.columns.{field}", default_value)
                    fixes_applied.append(f"✅ {field} automatisch gesetzt")
                    print(f"🛠️ Auto-fix: modulkombi.columns.{field} = '{default_value}'")
                
                # Problem 2: Wert existiert nicht in Excel
                elif config_value not in actual_columns and field != "Termine_Feld":
                    issues_found.append(f"Spalte '{config_value}' nicht in Excel gefunden")
                    print(f"❌ Config-Fehler: '{config_value}' existiert nicht in '{filename}'")
                    
                    # Finde ähnliche Spalten für Vorschlag
                    similar = [col for col in actual_columns if field.lower() in col.lower()]
                    if similar:
                        print(f"💡 Ähnliche Spalten gefunden: {similar}")
        
        except Exception as e:
            print(f"❌ Fehler beim Prüfen von Modulkombi: {e}")
    
    # ============================================================
    # 2. PRÜFE STATUS SELECT SHEET (für Termine)
    # ============================================================
    if filename and os.path.exists(filename):
        try:
            # Prüfe ob Sheet "Status Select" existiert
            wb = openpyxl.load_workbook(filename, read_only=True)
            if "Status Select" in wb.sheetnames:
                ws = wb["Status Select"]
                hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                
                print(f"📋 Status Select Sheet:")
                print(f"   Gefundene Spalten: {hdr}")
                
                # Wichtige Spalten prüfen
                if "Dauer" not in hdr:
                    issues_found.append("Sheet 'Status Select' fehlt Spalte 'Dauer'")
                else:
                    # KORRIGIERE Config für DauerFeld
                    if config_manager.get("modulkombi.columns.DauerFeld") != "Dauer":
                        config_manager.set("modulkombi.columns.DauerFeld", "Dauer in Tagen")
                        fixes_applied.append("✅ DauerFeld korrigiert auf 'Dauer'")
                        print("🛠️ Auto-fix: DauerFeld = 'Dauer' (für Status Select)")
                
                if "Termine" not in hdr:
                    issues_found.append("Sheet 'Status Select' fehlt Spalte 'Termine'")
                
                wb.close()
            else:
                issues_found.append("Sheet 'Status Select' existiert nicht")
                print("❌ Sheet 'Status Select' fehlt in der Excel-Datei")
        
        except Exception as e:
            print(f"❌ Fehler beim Prüfen von Status Select: {e}")
    
    # ============================================================
    # 3. ZUSAMMENFASSUNG
    # ============================================================
    if fixes_applied:
        print(f"\n🛠️ Automatische Fixes durchgeführt:")
        for fix in fixes_applied:
            print(f"   {fix}")
    
    if issues_found:
        print(f"\n❌ {len(issues_found)} kritische Probleme gefunden:")
        for issue in issues_found:
            print(f"   - {issue}")
        
        # Zeige dem Benutzer eine klare Meldung
        msg = "⚠️ Config-Probleme erkannt und teilweise automatisch behoben.\n\n"
        if fixes_applied:
            msg += "✅ Folgende Fixes wurden angewendet:\n" + "\n".join(fixes_applied) + "\n\n"
        msg += "Bitte überprüfen Sie die Log-Datei für Details."
        
        print("\n" + "="*70)
        print("CONFIG-DIAGNOSE-ERGEBNIS:")
        print("="*70)
        print(msg)
        print("="*70 + "\n")
        
        return len(issues_found) == 0  # True wenn keine kritischen Probleme
    
    print("✅ Config-Diagnose abgeschlossen - keine Probleme gefunden")
    return True
        

stop_event = threading.Event()
write_lock = threading.RLock()
_token_lock = threading.Lock()



os.environ["DEBUG_KEYWORD"] = "0" # für Debug der Dialogfelder auf '1' setzen


print("⏳ Lade FastText-Sprachmodell...")
def get_fasttext_model():
    # Wenn als EXE, dann liegt das Modell im _MEIPASS-Ordner
    if hasattr(sys, '_MEIPASS'):
        model_path = os.path.join(sys._MEIPASS, "lid.176.ftz")
    else:
        # Normaler Python Pfad
        model_path = os.path.join(os.path.dirname(__file__), "lid.176.ftz")
    return fasttext.load_model(model_path)

fasttext_model = get_fasttext_model()
print("⏳ Lade Spacy-Srachmodell...")
def load_spacy_model(name="de_core_news_md"):
    """Lädt spaCy-Modell - im Bundle OHNE Download"""
    
    # 🔥 IM BUNDLE: Kein Download, direkt laden oder sterben
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        try:
            print(f"📦 Lade Modell '{name}' aus Bundle...")
            return spacy.load(name, disable=["parser", "ner", "lemmatizer"])
        except OSError:
            print(f"❌ KRITISCH: Modell '{name}' fehlt im Bundle!")
            print("   → EXE ist defekt! Bitte neu bauen mit --collect-all")
            sys.exit(1)

    # 💻 ENTWICKLUNG: Normaler Modus
    try:
        print(f"📂 Lade Modell '{name}' (Development)...")
        return spacy.load(name, disable=["parser", "ner", "lemmatizer"])
    except OSError:
        print(f"⚠️  Modell '{name}' nicht gefunden.")
        print(f"   Installiere: python -m spacy download {name}")
        sys.exit(1)


nlp_de = load_spacy_model()

#print(spacy.util.get_package_path("de_core_news_md"))


CONFIG_FILE = "config.json"

class ConfigManager:
    def __init__(self):
        self.config = {
            "modulkombi": {"filename": None, "sheet": None, "columns": {}},
            "module": {"filename": None, "sheet": None, "columns": {}},
            "mapping": {"filename": None, "sheet": None, "columns": {}},
            "keywords": {"filename": None, "sheet": None, "columns": {}},
            "systematik": {"filename": None, "columns": {}},
            "prompts": {"filename": None, "columns": {}}
        }
        self.load()
    
    def load(self):
        """Lädt die komplette Config-Datei und ersetzt die aktuelle Struktur"""
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    loaded_config = json.load(f)
                    # VOLLSTÄNDIGE Ersetzung (wichtig!)
                    self.config = loaded_config
            except Exception as e:
                print(f"Warnung: Konnte Config nicht laden: {e}. Verwende Standard-Konfiguration.")
    
    def save(self):
        """Speichert die aktuelle Config-Struktur"""
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(self.config, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"Fehler beim Speichern der Config: {e}")
    
    def get(self, key: str, default=None) -> Any:
        """
        Holt einen Wert aus der verschachtelten Config mittels dotted notation.
        Beispiele:
        - get("modulkombi.filename") -> "C:/Pfad/...xlsx"
        - get("module.columns.Modulnummer1") -> "interne Angebotsnummer"
        """
        try:
            keys = key.split(".")
            value = self.config
            for k in keys:
                value = value[k]
            return value
        except (KeyError, TypeError, AttributeError):
            return default
    
    def set(self, key: str, value: Any):
        """
        Setzt einen Wert in der verschachtelten Config mittels dotted notation.
        Erstellt fehlende Strukturen automatisch.
        """
        keys = key.split(".")
        target = self.config
        
        # Durchlaufe alle Schlüssel außer den letzten
        for k in keys[:-1]:
            if k not in target:
                target[k] = {}
            target = target[k]
        
        # Setze den letzten Schlüssel
        target[keys[-1]] = value
        self.save()

# Globale Instanz
config_manager = ConfigManager()


class ColumnMappingDialog(QDialog):
    def __init__(self, file_type: str, df: pd.DataFrame, parent=None):
        super().__init__(parent)
        self.file_type = file_type
        self.df = df
        self.setWindowTitle(f"Spalten zuordnen: {file_type}")
        self.setMinimumWidth(700)
        
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(f"Ordne die benötigten Spalten für <b>{file_type}</b> den tatsächlichen Spalten in der Datei zu:"))
        
        # Welche Spalten werden benötigt?
        self.required_fields = {
            "modulkombi": ["Modulnummern", "Kursübersicht", "EinleitungNeu", "TitelFeld", "TitelNeu", 
                          "Systematik_Feld", "Keyword_Feld", "Termine_Feld", "Inhalte_Feld", "DauerFeld",
                          "Zielgruppe", "Voraussetzungen", "Abschlussart", "Abschlussbezeichnung"],
            "module": ["Modulnummer1", "Modulnummer2", "Titel", "Inhalt", "Beschreibung"],
            "mapping": ["Modulnummer Alt", "Modulnummer Neu", "Modulnummer HP"],
            "keywords": ["Kategorie Suchvolumen"]
        }.get(file_type, [])
        
        # Formular erstellen
        self.mapping_widgets = {}
        form_layout = QFormLayout()
        
        # Verfügbare Spalten aus dem DataFrame
        available_columns = [" "] + list(df.columns)
        
        for field in self.required_fields:
            # WICHTIG: Hole den STRING-Wert, nicht ein dict!
            columns = config_manager.get(f"{file_type}.columns", {})
            current_mapping = columns.get(field, "")  # <- DAS WAR DAS PROBLEM!
            
            combo = QComboBox()
            combo.addItems(available_columns)
            
            # Setze aktuelle Auswahl (falls vorhanden)
            if current_mapping:
                idx = combo.findText(current_mapping)
                if idx >= 0:
                    combo.setCurrentIndex(idx)
            
            self.mapping_widgets[field] = combo
            form_layout.addRow(f"{field}:", combo)
        
        layout.addLayout(form_layout)
        
        # Buttons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def get_mapping(self) -> dict:
        """Gibt die Zuordnung zurück als {feld: spaltenname}"""
        return {field: widget.currentText() for field, widget in self.mapping_widgets.items()}



client = OpenAI(
  api_key="insertAPIKey"
)


AnzahlKurse = 100  # Anzahl der Kurse, die mit KI überarbeitet werden sollen
MaxRows = 100 # Anzahl der Kurse die Überhaupt überarbeitet werden sollen
MAX_RETRIES = 2
WAIT_SECONDS = 10  # in Sekunden

MAX_TOKENS_PER_MIN = 22000



DEBUG = True

# =============
# Felder auswählen, die bearbeitet werden
# =============
@dataclass
class GenerationOptions:
    createMasterKeyword: bool = True
    createTitel: bool = True
    createEinleitung: bool = True
    createInhalte: bool = True
    createZielgruppe: bool = True
    createVoraussetzungen: bool = True
    createAbschlussart: bool = True
    createAbschlussbezeichnung: bool = True
    createSystematik: bool = True
    createKeywords: bool = True
    createTermine:bool = False



STOPWORDS: Set[str] = {
    "mit", "und", "oder", "für", "von", "im", "in", "auf", "am", "der",
    "die", "das", "and", "or", "for", "of", "the", "a", "an", "in", "on", "with", "by", "to", "online",
    "new", "digital", "certified", "certification", "good", "best"
    "ein", "eine", "einer", "einem", "eines", "dem", "den", "des",
    "neu", "neue", "neuer", "neues", "digital", "digitales", "digitale", 
    "digitaler", "agil", "agiler", "agile", "agiles", "agilen", "beste", "gute", "besser",
    "zertifiziert", "kommunikations", "kommunikation", "konfliktmanagement", 
    "zertifikate", "arbeiten", "zertifizierungen", "strategien", "kompetenz", "vertiefung", "zusammenarbeit",
    "zertifizierung", "zertifikat", "management", "fortbildung", "ausbildung", "hr weiterbildung", "karriere", "studium", "weiterbildung", "training", "schulung", "seminar", "prüfung", "it kurs"
}

_STOP_RE = re.compile(
    r"\b(?:" + "|".join(map(re.escape, STOPWORDS)) + r")\b",
    flags=re.IGNORECASE
)


SHORT_WORD_WHITELIST: Set[str] = {
    "sap", "aws", "sql", "c++", "c#", "ai", "itil", "scrum", "devops", "azure", "ki", "seo", "hana"
    }

_SHORT_RE = re.compile(
    r"\b(?:" + "|".join(map(re.escape, SHORT_WORD_WHITELIST)) + r")\b",
    flags=re.IGNORECASE
)

SPECIAL_WORDS: Dict[str, str] = {"ki": "KI", "ai": "AI", "azure": "Azure", "itil": "ITIL", "it": "IT", "sap": "SAP", "ms": "MS", "lpi": "LPI", "edv": "EDV",
                 "comptia": "CompTIA", "pcep": "PCEP", "we": "WE", "php": "PHP", "ihk": "IHK", "lpic": "LPIC", "mysql": "MySQL", "oca": "OCA", "pc": "PC",
                 "ooc": "OOC", "html": "HTML", "css": "CSS", "ui": "UI", "ux": "UX", "devops": "DevOps", "aevo": "AEVO", "tüv": "TÜV", "qmb": "QMB", "qm": "QM",
                 "ms": "MS"}

COMPOSITE_EXCEPTIONS = [
    "E-Commerce",
    "E-Mail",
    "E-Learning",
    "IT-Sicherheit",
    "IT-Service",
    "IT-Techniker",
    "Online-Marketing",
    "IT-Projektmanagement",
    "SCRUM-Master",
    "Qualitätsmanager",
    "ERP-Berater",
    "IT-Dispatcher",
    "IT-Ticket-System",
    "IT-Projektkoordination",
    "Qualitätsmanagement-Kenntnissen",
    "IT-Infrastrukturtechniker",
    "SEO-Manager",
    "Arbeiten 4.0",
    "IT Security",
    "IT-Projektassistenz",
    "ADA-Schein",
    "Microsoft-Systeme - Netzwerkadministration",
    "Büromanager",
    "Practitioner",
    "PC-Experte",
    "Software-Engineer",
    "UI/UX-Designer",
    "JavaScript",
    "IHK-Zertifikat",
    "Change",
    "IT-Bereich",
    "IHK-AEVO",
    "Teammanager-Projektorganisation",
    "IT-Projektleiter",
    "TÜV-zertifizierte",
    "ADA-IHK",
    "managen",
    "Office 4.0",
    "UC_FL",
    "IT-Projektleitung"
]

ENGLISH_TERMS = {
    "admin", "administrator", "agile", "ai", "analytics", "api",
    "artificial", "associate", "backend", "bi", "blog", "bootstrap",
    "brand", "browser", "business", "campaign", "certified", "change",
    "cloud", "cloud architect", "cms", "computer", "consultant",
    "content", "controller", "crm", "css", "customer", "data",
    "database", "design", "designer", "desktop", "devops",
    "ecommerce", "e-mail", "engineer", "enterprise", "erp",
    "ethical", "facebook", "frontend", "fullstack", "fundamentals",
    "game", "google", "governance", "html", "http", "hybrid",
    "hypertext", "identity", "infrastructure", "intelligence",
    "internet", "intranet", "java", "javascript", "jira", "kanban",
    "laravel", "leadership", "level", "linux", "login", "manager",
    "managing", "marketing", "markup", "media", "modern", "mysql",
    "network", "neuromarketing", "node", "nodejs", "online", "owner",
    "pc", "performance", "php", "planning", "platform", "power",
    "powershell", "process", "product", "professional", "programmer",
    "programming", "project", "projects", "python", "quality",
    "react", "reporting", "request", "responsive", "sales",
    "scrum", "security", "server", "service", "shop", "social",
    "software", "specialist", "sql", "ssl", "stack", "strategic",
    "support", "system", "systems", "tag", "team", "technical",
    "token", "tools", "traffic", "trainer", "training",
    "transformation", "ui", "url", "user", "ux", "version",
    "virtual", "vpn", "warehouse", "web", "webanalytics",
    "webdesign", "webdesigner", "webdeveloper", "webdevelopment",
    "webmaster", "website", "webshop", "whitehat", "windows",
    "wordpress", "workflow", "xpert", "youtube"
}

_COMPOSITES_RE = re.compile(
    "|".join(map(re.escape, COMPOSITE_EXCEPTIONS)),
    flags=re.IGNORECASE
)


# Variablen für Variations Excel
Modulnummern = 'Modulnummern kommagetrennt NEU\nNach Übernahme '
Kursübersicht = 'Kursübersicht der Einzelmodule (NEU)'
EinleitungNeu = 'Einleitung NEU Sabine vgl. Spalte I\n\nGrün: in ESTHER übernommen'
TitelFeld = 'Neu zertifizierter Titel der Maßnahme NEU'
TitelNeu = 'Titeloptimierung Edtelligent/ Sabine\nvgl. Spalte C rosa\nVgl. Bericht 5.8\nGrün: in ESTHER übernommen'
Systematik_Feld = 'Systematik in Kursnet'
Keyword_Feld = 'Keyword-Optimierung Sabine/ Edtelligent 04_25\n\nVgl. Bericht 5.6\nGrün: in ESTHER übernommen'
Termine_Feld = 'mögliche Zeiträume für ESTHER \n(rot = keine Kombination gefunden, \ngrün = fertig) '
Inhalte_Feld = 'Inhalt NEU'
DauerFeld = 'Dauer in Tagen'

# -------------------
# Pflichtspalten für modulkombi_df (werden geprüft)
# -------------------
REQUIRED_COLUMNS = [
    "Modulnummern kommagetrennt NEU\nNach Übernahme ",
    "Kursübersicht der Einzelmodule (NEU)",
    "Einleitung NEU Sabine vgl. Spalte I\n\nGrün: in ESTHER übernommen",
    "Neu zertifizierter Titel der Maßnahme NEU",
    "Titeloptimierung Edtelligent/ Sabine\nvgl. Spalte C rosa\nVgl. Bericht 5.8\nGrün: in ESTHER übernommen",
    "Systematik in Kursnet",
    "Keyword-Optimierung Sabine/ Edtelligent 04_25\n\nVgl. Bericht 5.6\nGrün: in ESTHER übernommen",
    "mögliche Zeiträume für ESTHER \n(rot = keine Kombination gefunden, \ngrün = fertig) ",
    "Inhalt NEU",
    "Dauer in Tagen",
    "Zielgruppe",
    "Voraussetzungen",
    "Abschlussart",
    "Abschlussbezeichnung",
]

# Pflichtspalten für module_df (Kursinhalte)
REQUIRED_MODULE_COLUMNS = ["Modulnummer1", "Modulnummer2", "Titel", "Inhalt"]

# Pflichtspalten für mapping_df (Mapping)
REQUIRED_MAPPING_COLUMNS = ["Angebotsnummer ALT", "Angebotsnummer NEU", "Modulnummer HP"]

# -------------------
# Globale Variablen (werden vom Loader gesetzt)
# -------------------
modulkombi_df = None
module_df = None
mapping_df = None
keywords_df = None
volume_map = None
prompts = None
SYSTEMATIK = None
USE_AZURE_WRITE = False

# Zurückgelegte Dateiinformationen (zu Debug/Logging)
FILENAME = None
SHEETNAME = None
OTHER_FILES = {}




# -------------------
# Hilfsdialog zum Eingeben von Azure-/SharePoint-Parametern
# -------------------
class AzureParamsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Azure / SharePoint Parameter")
        self.setMinimumWidth(480)
        layout = QVBoxLayout(self)

        form = QFormLayout()
        self.tenant_edit = QLineEdit()
        self.client_edit = QLineEdit()
        self.secret_edit = QLineEdit()
        self.secret_edit.setEchoMode(QLineEdit.Password)

        self.sharepoint_site = QLineEdit()
        self.site_name = QLineEdit()
        self.drive_name = QLineEdit()
        self.file_path = QLineEdit()
        self.sheet_name = QLineEdit()

        form.addRow("TENANT_ID:", self.tenant_edit)
        form.addRow("CLIENT_ID:", self.client_edit)
        form.addRow("CLIENT_SECRET:", self.secret_edit)
        form.addRow("SHAREPOINT_SITE (z.B. contoso.sharepoint.com):", self.sharepoint_site)
        form.addRow("SITE_NAME (Pfad unter Site):", self.site_name)
        form.addRow("DRIVE_NAME:", self.drive_name)
        form.addRow("FILE_PATH (z.B. folder/file.xlsx):", self.file_path)
        form.addRow("BlattName:", self.sheet_name)

        layout.addLayout(form)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def get_params(self) -> Optional[Dict[str, str]]:
        if self.exec() == QDialog.Accepted:
            return {
                "TENANT_ID": self.tenant_edit.text().strip(),
                "CLIENT_ID": self.client_edit.text().strip(),
                "CLIENT_SECRET": self.secret_edit.text().strip(),
                "SHAREPOINT_SITE": self.sharepoint_site.text().strip(),
                "SITE_NAME": self.site_name.text().strip(),
                "DRIVE_NAME": self.drive_name.text().strip(),
                "FILE_PATH": self.file_path.text().strip(),
                "BlattName": self.sheet_name.text().strip(),
            }
        return None


# -------------------
# Funktion: Auth + Download via Graph (Drive item content)
# -------------------
def download_file_from_sharepoint(params: Dict[str, str]) -> bytes:
    """
    Erwartet params mit TENANT_ID, CLIENT_ID, CLIENT_SECRET, SHAREPOINT_SITE, SITE_NAME, DRIVE_NAME, FILE_PATH
    Liefert: Datei-Bytes (z.B. Excel-Bytes oder JSON-Bytes)
    """
    TENANT_ID = params["TENANT_ID"]
    CLIENT_ID = params["CLIENT_ID"]
    CLIENT_SECRET = params["CLIENT_SECRET"]
    SHAREPOINT_SITE = params["SHAREPOINT_SITE"]
    SITE_NAME = params["SITE_NAME"]
    DRIVE_NAME = params["DRIVE_NAME"]
    FILE_PATH = params["FILE_PATH"]

    AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
    SCOPES = ["https://graph.microsoft.com/.default"]

    app = ConfidentialClientApplication(CLIENT_ID, authority=AUTHORITY, client_credential=CLIENT_SECRET)
    result = app.acquire_token_silent(SCOPES, account=None)
    if not result:
        result = app.acquire_token_for_client(scopes=SCOPES)

    if "access_token" not in result:
        raise Exception("Fehler bei der Authentifizierung: " + str(result.get("error_description", result)))

    headers = {"Authorization": "Bearer " + result["access_token"]}

    # Site abrufen
    site_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE}:/{SITE_NAME}"
    site = requests.get(site_url, headers=headers).json()
    if "id" not in site:
        raise Exception(f"Fehler beim Abrufen der Site: {site}")
    site_id = site["id"]

    # Drive ID
    drive_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
    drives = requests.get(drive_url, headers=headers).json()
    if "value" not in drives:
        raise Exception(f"Fehler beim Abrufen der Drives: {drives}")
    drive_id = next((d["id"] for d in drives["value"] if d.get("name") == DRIVE_NAME), None)
    if drive_id is None:
        raise Exception(f"Drive '{DRIVE_NAME}' nicht gefunden. Verfügbare drives: {[d.get('name') for d in drives.get('value', [])]}")

    # Datei-Item
    file_item_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}/root:/{FILE_PATH}"
    file_item = requests.get(file_item_url, headers=headers).json()
    if "id" not in file_item:
        raise Exception(f"Fehler beim Abrufen des Datei-Items: {file_item}")
    item_id = file_item["id"]

    # download content endpoint
    download_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content"
    resp = requests.get(download_url, headers=headers, stream=True)
    if resp.status_code not in (200, 201):
        raise Exception(f"Fehler beim Download ({resp.status_code}): {resp.text}")

    return resp.content


# -------------------
# Hilfs-Funktion: Prüfen und ggf. Spaltenmapping anfordern
# -------------------
def ensure_modulkombi_columns(parent: QWidget, df: pd.DataFrame) -> Tuple[bool, pd.DataFrame]:
    """
    Prüft, ob alle REQUIRED_COLUMNS in df existieren.
    Wenn nicht, fragt den Nutzer interaktiv per Dialog für jedes fehlende Feld, welches vorhandene Feld gemappt werden soll.
    Falls der Nutzer Abbruch wählt -> returns (False, df)
    Wenn Mapping erfolgreich -> df mit umbenannten Spalten zurückgeben und True
    """
    existing = set(df.columns.tolist())
    missing = [c for c in REQUIRED_COLUMNS if c not in existing]

    if not missing:
        # alles vorhanden
        return True, df

    # Benutzer informieren
    QMessageBox.warning(parent, "Fehlende Spalten",
                        "Es fehlen Pflichtspalten in der geladenen Datei. Du wirst nun für jede fehlende Pflichtspalte aufgefordert, eine vorhandene Spalte zuzuordnen. "
                        "Wenn du abbrechen möchtest, wähle 'Abbrechen' beim Mapping-Dialog und wähle die Datei erneut.")

    new_names = {}
    available = ["<Datei neu wählen>"] + df.columns.tolist()

    for req in missing:
        item, ok = QInputDialog.getItem(parent, "Spalten-Mapping",
                                       f"Bitte wähle die vorhandene Spalte, die zur Pflichtspalte\n\n'{req}'\n\npasst (oder wähle '<Datei neu wählen>' um eine andere Datei auszuwählen):",
                                       available, 0, False)
        if not ok:
            # Abbruch durch Benutzer
            return False, df
        if item == "<Datei neu wählen>":
            return False, df
        # mappe ausgewählte vorhandene Spalte auf required name
        new_names[item] = req

    # benenne Spalten um: keys = exist col, values = required name
    df = df.rename(columns=new_names)
    # nochmal prüfen komplettheit
    remaining_missing = [c for c in REQUIRED_COLUMNS if c not in df.columns.tolist()]
    if remaining_missing:
        QMessageBox.critical(parent, "Mapping fehlerhaft", f"Nach dem Mapping fehlen noch Spalten: {remaining_missing}. Datei neu auswählen.")
        return False, df

    return True, df

def ensure_module_columns(parent: QWidget, df: pd.DataFrame) -> Tuple[bool, pd.DataFrame]:
    """Prüft ob alle REQUIRED_MODULE_COLUMNS existieren, lässt bei Bedarf Mapping zu."""
    existing = set(df.columns.tolist())
    missing = [c for c in REQUIRED_MODULE_COLUMNS if c not in existing]
    
    if not missing:
        return True, df
        
    QMessageBox.warning(parent, "Fehlende Spalten in Kursinhalten",
                       "Es fehlen Pflichtspalten in der geladenen Datei. Du wirst nun für jede fehlende Pflichtspalte aufgefordert, eine vorhandene Spalte zuzuordnen.")
    
    new_names = {}
    available = ["<Datei neu wählen>"] + df.columns.tolist()
    
    for req in missing:
        item, ok = QInputDialog.getItem(parent, "Spalten-Mapping Kursinhalte",
                                       f"Ordne die Pflichtspalte\n\n'{req}'\n\neiner vorhandenen Spalte zu:",
                                       available, 0, False)
        if not ok or item == "<Datei neu wählen>":
            return False, df
        new_names[item] = req
    
    df = df.rename(columns=new_names)
    remaining = [c for c in REQUIRED_MODULE_COLUMNS if c not in df.columns]
    if remaining:
        QMessageBox.critical(parent, "Mapping fehlerhaft", f"Fehlen noch: {remaining}")
        return False, df
        
    return True, df

def ensure_mapping_columns(parent: QWidget, df: pd.DataFrame) -> Tuple[bool, pd.DataFrame]:
    """Prüft ob alle REQUIRED_MAPPING_COLUMNS existieren, lässt bei Bedarf Mapping zu."""
    existing = set(df.columns.tolist())
    missing = [c for c in REQUIRED_MAPPING_COLUMNS if c not in existing]
    
    if not missing:
        return True, df
        
    QMessageBox.warning(parent, "Fehlende Spalten in Mapping",
                       "Es fehlen Pflichtspalten in der geladenen Datei. Du wirst nun für jede fehlende Pflichtspalte aufgefordert, eine vorhandene Spalte zuzuordnen.")
    
    new_names = {}
    available = ["<Datei neu wählen>"] + df.columns.tolist()
    
    for req in missing:
        item, ok = QInputDialog.getItem(parent, "Spalten-Mapping Mapping-Datei",
                                       f"Ordne die Pflichtspalte\n\n'{req}'\n\neiner vorhandenen Spalte zu:",
                                       available, 0, False)
        if not ok or item == "<Datei neu wählen>":
            return False, df
        new_names[item] = req
    
    df = df.rename(columns=new_names)
    remaining = [c for c in REQUIRED_MAPPING_COLUMNS if c not in df.columns]
    if remaining:
        QMessageBox.critical(parent, "Mapping fehlerhaft", f"Fehlen noch: {remaining}")
        return False, df
        
    return True, df

class ColumnMappingSetupDialog(QDialog):
    """Dialog zum manuellen Eintragen aller Spaltennamen ohne Excel-Datei"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Spaltennamen manuell konfigurieren")
        self.setMinimumWidth(700)
        
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Gib die exakten Spaltennamen aus deinen Excel-Dateien ein:"))
        
        # Tabs für verschiedene Dateitypen
        tabs = QTabWidget()
        self.fields = {}
        
        # Feld-Definitionen mit Standardwerten (Fallback, falls Config leer ist)
        field_defaults = {
            "modulkombi": {
                "Modulnummern": "Modulnummern kommagetrennt NEU\nNach Übernahme ",
                "Kursübersicht": "Kursübersicht der Einzelmodule (NEU)",
                "EinleitungNeu": "Einleitung NEU Sabine vgl. Spalte I\n\nGrün: in ESTHER übernommen",
                "TitelFeld": "Neu zertifizierter Titel der Maßnahme NEU",
                "TitelNeu": "Titeloptimierung Edtelligent/ Sabine\nvgl. Spalte C rosa\nVgl. Bericht 5.8\nGrün: in ESTHER übernommen",
                "Systematik_Feld": "Systematik in Kursnet",
                "Keyword_Feld": "Keyword-Optimierung Sabine/ Edtelligent 04_25\n\nVgl. Bericht 5.6\nGrün: in ESTHER übernommen",
                "Termine_Feld": "mögliche Zeiträume für ESTHER \n(rot = keine Kombination gefunden, \ngrün = fertig) ",
                "Inhalte_Feld": "Inhalt NEU",
                "DauerFeld": "Dauer in Tagen",
                "Zielgruppe": "Zielgruppe",
                "Voraussetzungen": "Voraussetzungen",
                "Abschlussart": "Abschlussart",
                "Abschlussbezeichnung": "Abschlussbezeichnung"
            },
            "module": {
                "Modulnummer1": "Modulnummer1",
                "Modulnummer2": "Modulnummer2",
                "Titel": "Titel",
                "Inhalt": "Inhalt",
                "Beschreibung": "Beschreibung"
            },
            "mapping": {
                "Modulnummer Alt": "Modulnummer Alt",
                "Modulnummer Neu": "Modulnummer Neu",
                "Modulnummer HP": "Modulnummer HP"
            },
            "keywords": {
                "Kategorie Suchvolumen": "Kategorie Suchvolumen"
            }
        }
        
        for file_type in ["modulkombi", "module", "mapping", "keywords"]:
            tab = QWidget()
            tab_layout = QFormLayout(tab)
            
            for field, default_value in field_defaults.get(file_type, {}).items():
                # Hole aktuellen Wert aus Config oder verwende Default
                current_value = config_manager.get(f"{file_type}.columns.{field}", default_value)
                
                line_edit = QLineEdit(current_value)
                line_edit.setPlaceholderText(f"Spaltenname für {field}")
                self.fields[f"{file_type}.{field}"] = line_edit
                tab_layout.addRow(f"{field}:", line_edit)
            
            tabs.addTab(tab, file_type.title())
        
        layout.addWidget(tabs)
        
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def accept(self):
        """Speichert die neuen Werte SICHER"""
        try:
            for key, widget in self.fields.items():
                file_type, field = key.split(".", 1)
                
                # SICHERHEIT: Prüfe und erstelle die Struktur falls nötig
                if file_type not in config_manager.config:
                    config_manager.config[file_type] = {"filename": None, "sheet": None, "columns": {}}
                if "columns" not in config_manager.config[file_type]:
                    config_manager.config[file_type]["columns"] = {}
                
                # Setze den Wert
                config_manager.config[file_type]["columns"][field] = widget.text()
            
            # Speichern
            config_manager.save()
            
            # Erfolgsmeldung
            QMessageBox.information(self, "Erfolg", 
                                  "Spaltennamen wurden erfolgreich gespeichert!\n\n"
                                  "Die Änderungen werden beim nächsten Programm-Start aktiv.")
            
            # Globale Variablen neu initialisieren
            initialize_column_config()
            
            super().accept()
            
        except Exception as e:
            QMessageBox.critical(self, "Fehler beim Speichern", 
                               f"Die Spaltennamen konnten nicht gespeichert werden:\n\n{e}")
        
# Dynamische Spalten-Variablen (werden beim Start gesetzt)
def initialize_column_config():
    """Initialisiert alle globalen Spaltennamen-Variablen aus der Config"""
    global Modulnummern, Kursübersicht, EinleitungNeu, TitelFeld, TitelNeu
    global Systematik_Feld, Keyword_Feld, Termine_Feld, Inhalte_Feld, DauerFeld
    global REQUIRED_COLUMNS, REQUIRED_MODULE_COLUMNS, REQUIRED_MAPPING_COLUMNS
    
    def get_safe_config(key: str, default: str) -> str:
        """Holt Config-Wert, Fallback bei Leerstring."""
        value = config_manager.get(key, default)
        if not value or value.strip() == "":
            print(f"⚠️ Leerer Config-Wert für '{key}', verwende Fallback")
            return default
        return value
    
    Modulnummern = get_safe_config("modulkombi.columns.Modulnummern", 
                                   "Modulnummern kommagetrennt NEU\nNach Übernahme ")
    Kursübersicht = get_safe_config("modulkombi.columns.Kursübersicht",
                                   "Kursübersicht der Einzelmodule (NEU)")
    EinleitungNeu = get_safe_config("modulkombi.columns.EinleitungNeu",
                                   "Einleitung NEU Sabine vgl. Spalte I\n\nGrün: in ESTHER übernommen")
    TitelFeld = get_safe_config("modulkombi.columns.TitelFeld",
                               "Neu zertifizierter Titel der Maßnahme NEU")
    TitelNeu = get_safe_config("modulkombi.columns.TitelNeu",
                              "Titeloptimierung Edtelligent/ Sabine\nvgl. Spalte C rosa\nVgl. Bericht 5.8\nGrün: in ESTHER übernommen")
    Systematik_Feld = get_safe_config("modulkombi.columns.Systematik_Feld",
                                     "Systematik in Kursnet")
    Keyword_Feld = get_safe_config("modulkombi.columns.Keyword_Feld",
                                  "Keyword-Optimierung Sabine/ Edtelligent 04_25\n\nVgl. Bericht 5.6\nGrün: in ESTHER übernommen")
    
    # WICHTIG: Für 'Status Select' Sheet verwenden wir EINFACHE Namen!
    Termine_Feld = "Termine"  # HARDCODIERT - dies ist der echte Excel-Header
    
    Inhalte_Feld = get_safe_config("modulkombi.columns.Inhalte_Feld", "Inhalt NEU")
    
    # DauerFeld wird automatisch angepasst basierend auf Sheet-Struktur
    DauerFeld = get_safe_config("modulkombi.columns.DauerFeld", "Dauer in Tagen")
    
    # ============================================================
    # PRÜFE OB 'Status Select' existiert und korrigiere DauerFeld
    # ============================================================
    filename = config_manager.get("modulkombi.filename")
    if filename and os.path.exists(filename):
        try:
            wb = openpyxl.load_workbook(filename, read_only=True)
            if "Status Select" in wb.sheetnames:
                ws = wb["Status Select"]
                hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                if "Dauer" in hdr:
                    DauerFeld = "Dauer"  # KORRIGIERE für Status Select
                    print("✅ DauerFeld automatisch korrigiert zu 'Dauer' (Status Select)")
            wb.close()
        except:
            pass  # Fehler ignorieren, ist nicht kritisch
    
    # ============================================================
    # PFLICHTSPALTEN LISTE ERSTELLEN
    # ============================================================
    REQUIRED_COLUMNS = [
        Modulnummern, Kursübersicht, EinleitungNeu, TitelFeld, TitelNeu,
        Systematik_Feld, Keyword_Feld, Termine_Feld, Inhalte_Feld, DauerFeld,
        "Zielgruppe", "Voraussetzungen", "Abschlussart", "Abschlussbezeichnung"
    ]
    
    # ============================================================
    # LOGGING
    # ============================================================
    print("✅ Spaltennamen initialisiert:")
    print(f"   DauerFeld: '{DauerFeld}'")
    print(f"   Termine_Feld: '{Termine_Feld}'")
    print(f"   TitelFeld: '{TitelFeld}'")
    print(f"   Anzahl Pflichtspalten: {len(REQUIRED_COLUMNS)}")
    
    # Module Spalten (Fallback auf Standardwerte)
    REQUIRED_MODULE_COLUMNS = [
        config_manager.get("module.columns.Modulnummer1", "Modulnummer1"),
        config_manager.get("module.columns.Modulnummer2", "Modulnummer2"),
        config_manager.get("module.columns.Titel", "Titel"),
        config_manager.get("module.columns.Inhalt", "Inhalt"),
        config_manager.get("module.columns.Beschreibung", "Beschreibung")
    ]
    
    # Mapping Spalten (Fallback auf Standardwerte)
    REQUIRED_MAPPING_COLUMNS = [
        config_manager.get("mapping.columns.Modulnummer Alt", "Modulnummer Alt"),
        config_manager.get("mapping.columns.Modulnummer Neu", "Modulnummer Neu"),
        config_manager.get("mapping.columns.Modulnummer HP", "Modulnummer HP")
    ]
    
    # Modulkombi Pflichtspalten
    REQUIRED_COLUMNS = [
        Modulnummern, Kursübersicht, EinleitungNeu, TitelFeld, TitelNeu,
        Systematik_Feld, Keyword_Feld, Termine_Feld, Inhalte_Feld, DauerFeld,
        "Zielgruppe", "Voraussetzungen", "Abschlussart", "Abschlussbezeichnung"
    ]



# -------------------
# Der Haupt-Dialog
# -------------------
class ExcelLoader(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Excel Multi-Loader")
        self.setMinimumWidth(800)  # Größer für Status-Anzeige

        self.main_layout = QVBoxLayout(self)
        
        # Config-Button ganz oben
        config_row = QHBoxLayout()
        self.btn_load_config = QPushButton("📂 Gesamte Config-Datei laden")
        self.btn_load_config.clicked.connect(self.load_full_config)
        config_row.addWidget(self.btn_load_config)
        
        self.btn_save_config = QPushButton("💾 Config speichern")
        self.btn_save_config.clicked.connect(self.save_full_config)
        config_row.addWidget(self.btn_save_config)
        
        self.btn_edit_columns = QPushButton("⚙️ Spaltennamen ändern")
        self.btn_edit_columns.clicked.connect(self.edit_column_names)
        config_row.addWidget(self.btn_edit_columns)
        
        self.main_layout.addLayout(config_row)
        self.main_layout.addWidget(QHLine())  # Horizontaler Trenner
        
        # Restlicher Inhalt
        self.main_layout.addWidget(QLabel("Bitte lade nacheinander alle erforderlichen Dateien:"))
        
        # self.azure_write = QCheckBox("🔄 Änderungen direkt in Azure speichern (nicht lokal)")
        # self.main_layout.addWidget(self.azure_write)
        
        # Status-Widgets
        self._status_widgets = {}
        self._loaded = {key: False for key in ["modulkombi", "module", "mapping", "keywords", "systematik", "prompts"]}
        
        # Datei-Zeilen hinzufügen
        self._add_file_row("modulkombi", "📊 Variationen (modulkombi_df)", self.load_modulkombi)
        self._add_file_row("module", "📚 Kursinhalte (module_df)", self.load_module_df)
        self._add_file_row("mapping", "🔀 Mapping (mapping_df)", self.load_mapping_df)
        self._add_file_row("keywords", "🔑 Keywords (keywords_df)", self.load_keywords_df)
        self._add_file_row("systematik", "📋 Systematiken", self.load_systematik)
        self._add_file_row("prompts", "📝 Prompts", self.load_prompts)
        
        # Buttons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.on_ok)
        buttons.rejected.connect(self.reject)
        self.main_layout.addWidget(buttons)

    def _add_file_row(self, key: str, label: str, load_func):
        """Erstellt eine Zeile mit Load-Button, Spalten-Button und Status"""
        row = QHBoxLayout()
        
        btn_load = QPushButton(label)
        btn_load.clicked.connect(load_func)
        row.addWidget(btn_load)
        
        # Spalten-Zuordnungs-Button NUR für Excel-Dateien (nicht für JSON)
        if key not in ["systematik", "prompts"]:  # 🚀 NEU: Kein Button für JSON-Dateien
            btn_map = QPushButton("🔄 Spalten")
            btn_map.clicked.connect(lambda checked, k=key: self.map_columns_for(k))
            btn_map.setToolTip(f"Spaltennamen für {key} zuordnen")
            row.addWidget(btn_map)
        
        # Status
        self._status_widgets[key] = QLabel("❌ Nicht geladen")
        row.addWidget(self._status_widgets[key])
        
        self.main_layout.addLayout(row)

    def load_full_config(self):
        """Lädt eine vollständige Config-Datei mit Pfaden und Spaltennamen"""
        fname, _ = QFileDialog.getOpenFileName(self, "Config-Datei laden", "", "JSON (*.json)")
        if not fname:
            return
        
        try:
            with open(fname, "r", encoding="utf-8") as f:
                new_config = json.load(f)
            
            # Validieren
            if not isinstance(new_config, dict) or "modulkombi" not in new_config:
                raise ValueError("Ungültiges Config-Format!")
            
            # Config laden
            config_manager.config = new_config
            config_manager.save()
            
            # Automatisch alle Dateien laden versuchen
            self._auto_load_all_files()
            
            QMessageBox.information(self, "Erfolg", "Config erfolgreich geladen! Die Dateien werden nun automatisch geladen.")
            
        except Exception as e:
            QMessageBox.critical(self, "Fehler", f"Config konnte nicht geladen werden:\n{e}")
            
            # DEBUG: Zeige geladene Config
        print("=== GELADENE CONFIG ===")
        print(json.dumps(config_manager.config, indent=2, ensure_ascii=False))
        
        # Prüfe, ob Pfade existieren
        for key in ["modulkombi", "module", "mapping"]:
            path = config_manager.get(f"{key}.filename")
            if path and os.path.exists(path):
                print(f"✅ {key}: Datei existiert - {path}")
            elif path:
                print(f"❌ {key}: Datei NICHT gefunden - {path}")
            else:
                print(f"⚠️ {key}: Kein Pfad in Config")
    
    def _auto_load_all_files(self):
        """Lädt ALLE Dateien aus der Config mit korrekter Fehlerbehandlung"""
        success_count = 0
        errors = []
        
        # Lokale Referenz auf globale Variablen
        global modulkombi_df, module_df, mapping_df, keywords_df, SYSTEMATIK, prompts
        global FILENAME, SHEETNAME, OTHER_FILES, volume_map
        
        # Stelle sicher, dass OTHER_FILES existiert
        if "OTHER_FILES" not in globals():
            globals()["OTHER_FILES"] = {}
        
        print("\n" + "="*60)
        print("AUTO-LOAD: Versuche Dateien aus Config zu laden...")
        print("="*60)
        
        # 1. Modulkombi laden
        filename = config_manager.get("modulkombi.filename")
        if filename and os.path.exists(filename):
            try:
                print(f"📊 Lade Modulkombi: {filename}")
                df = pd.read_excel(filename, sheet_name=config_manager.get("modulkombi.sheet"))
                
                # Spalten-Mapping anwenden
                columns = config_manager.get("modulkombi.columns", {})
                if columns:
                    reverse_mapping = {v: k for k, v in columns.items()}
                    df = df.rename(columns=reverse_mapping)
                
                # Global zuweisen
                globals()["modulkombi_df"] = df
                globals()["FILENAME"] = filename
                globals()["SHEETNAME"] = config_manager.get("modulkombi.sheet")
                globals()["OTHER_FILES"]["modulkombi"] = {
                    "filename": filename, 
                    "sheet": config_manager.get("modulkombi.sheet")
                }
                
                self._loaded["modulkombi"] = True
                self._status_widgets["modulkombi"].setText("✅ Auto-geladen")
                success_count += 1
                
            except Exception as e:
                errors.append(f"Modulkombi: {str(e)}")
                print(f"❌ Modulkombi-Fehler: {e}")
        elif filename:
            errors.append(f"Modulkombi: Datei nicht gefunden - {filename}")
        else:
            print("⚠️ Modulkombi: Kein Pfad in Config")
            
        
         # 2. Module laden
        filename = config_manager.get("module.filename")
        if filename and os.path.exists(filename):
            try:
                print(f"📚 Lade Module: {filename}")
                df = pd.read_excel(filename, sheet_name=config_manager.get("module.sheet"))
                
                # Spalten-Mapping anwenden
                columns = config_manager.get("module.columns", {})
                if columns:
                    reverse_mapping = {v: k for k, v in columns.items()}
                    df = df.rename(columns=reverse_mapping)
                
                # GLOBAL zuweisen - WICHTIG!
                globals()["module_df"] = df
                globals()["OTHER_FILES"]["module"] = {
                    "filename": filename, 
                    "sheet": config_manager.get("module.sheet")
                }
                
                self._loaded["module"] = True
                self._status_widgets["module"].setText("✅ Auto-geladen")
                success_count += 1
                
            except Exception as e:
                errors.append(f"Module: {str(e)}")
                print(f"❌ Module-Fehler: {e}")
        elif filename:
            errors.append(f"Module: Datei nicht gefunden - {filename}")
        else:
            print("⚠️ Module: Kein Pfad in Config")

        # 3. Mapping laden
        filename = config_manager.get("mapping.filename")
        if filename and os.path.exists(filename):
            try:
                print(f"🔀 Lade Mapping: {filename}")
                df = pd.read_excel(filename, sheet_name=config_manager.get("mapping.sheet"))
                
                columns = config_manager.get("mapping.columns", {})
                if columns:
                    reverse_mapping = {v: k for k, v in columns.items()}
                    df = df.rename(columns=reverse_mapping)
                
                globals()["mapping_df"] = df
                globals()["OTHER_FILES"]["mapping"] = {
                    "filename": filename, 
                    "sheet": config_manager.get("mapping.sheet")
                }
                
                self._loaded["mapping"] = True
                self._status_widgets["mapping"].setText("✅ Auto-geladen")
                success_count += 1
                
            except Exception as e:
                errors.append(f"Mapping: {str(e)}")
                print(f"❌ Mapping-Fehler: {e}")
        elif filename:
            errors.append(f"Mapping: Datei nicht gefunden - {filename}")
        else:
            print("⚠️ Mapping: Kein Pfad in Config")

        # 4. Keywords laden
        filename = config_manager.get("keywords.filename")
        if filename and os.path.exists(filename):
            try:
                print(f"🔑 Lade Keywords: {filename}")
                df = pd.read_excel(filename, sheet_name=config_manager.get("keywords.sheet"))
                
                # Volume Score berechnen
                globals()["volume_map"] = {'High': 3, 'Medium': 2, 'Low': 1}
                
                if 'Kategorie Suchvolumen' in df.columns:
                    df['volume_score'] = df['Kategorie Suchvolumen'].map(globals()["volume_map"])
                
                globals()["keywords_df"] = df
                globals()["OTHER_FILES"]["keywords"] = {
                    "filename": filename, 
                    "sheet": config_manager.get("keywords.sheet")
                }
                
                self._loaded["keywords"] = True
                self._status_widgets["keywords"].setText("✅ Auto-geladen")
                success_count += 1
                
            except Exception as e:
                errors.append(f"Keywords: {str(e)}")
                print(f"❌ Keywords-Fehler: {e}")
        elif filename:
            errors.append(f"Keywords: Datei nicht gefunden - {filename}")
        else:
            print("⚠️ Keywords: Kein Pfad in Config")

        # 5. Systematik laden
        filename = config_manager.get("systematik.filename")
        if filename and os.path.exists(filename):
            try:
                print(f"📋 Lade Systematik: {filename}")
                text = Path(filename).read_text(encoding="utf-8")
                globals()["SYSTEMATIK"] = json.loads(text)
                
                globals()["OTHER_FILES"]["systematik"] = {
                    "filename": filename
                }
                
                self._loaded["systematik"] = True
                self._status_widgets["systematik"].setText("✅ Auto-geladen")
                success_count += 1
                
            except Exception as e:
                errors.append(f"Systematik: {str(e)}")
                print(f"❌ Systematik-Fehler: {e}")
        elif filename:
            errors.append(f"Systematik: Datei nicht gefunden - {filename}")
        else:
            print("⚠️ Systematik: Kein Pfad in Config")

        # 6. Prompts laden
        filename = config_manager.get("prompts.filename")
        if filename and os.path.exists(filename):
            try:
                print(f"📝 Lade Prompts: {filename}")
                text = Path(filename).read_text(encoding="utf-8")
                globals()["prompts"] = json.loads(text)
                
                globals()["OTHER_FILES"]["prompts"] = {
                    "filename": filename
                }
                
                self._loaded["prompts"] = True
                self._status_widgets["prompts"].setText("✅ Auto-geladen")
                success_count += 1
                
            except Exception as e:
                errors.append(f"Prompts: {str(e)}")
                print(f"❌ Prompts-Fehler: {e}")
        elif filename:
            errors.append(f"Prompts: Datei nicht gefunden - {filename}")
        else:
            print("⚠️ Prompts: Kein Pfad in Config")
    
    def save_full_config(self):
        """Speichert die aktuell geladenen Pfade und Spaltennamen als vollständige Config-Datei"""
        
        # Sammle systematisch alle Daten
        config = {}
        errors = []
        
        # Debug-Ausgabe in Konsole
        print("\n" + "="*50)
        print("SAVE FULL CONFIG: Sammle aktuelle Datei-Informationen...")
        print("="*50)
        
        # === 1. Modulkombi ===
        try:
            modulkombi_data = {
                "filename": globals().get("FILENAME"),
                "sheet": globals().get("SHEETNAME"),
                "columns": config_manager.config.get("modulkombi", {}).get("columns", {})
            }
            config["modulkombi"] = modulkombi_data
            
            path = modulkombi_data["filename"] or "Nicht geladen"
            print(f"✅ Modulkombi: {path}")
            
        except Exception as e:
            errors.append(f"Modulkombi: {str(e)}")
            config["modulkombi"] = {"filename": None, "sheet": None, "columns": {}}
            print(f"❌ Modulkombi-Fehler: {e}")
        
        # === 2. Module ===
        try:
            other_files = globals().get("OTHER_FILES", {})
            module_file = other_files.get("module", {})
            
            module_data = {
                "filename": module_file.get("filename"),
                "sheet": module_file.get("sheet"),
                "azure_params": module_file.get("azure_params"),
                "columns": config_manager.config.get("module", {}).get("columns", {})
            }
            config["module"] = module_data
            
            # Zeige Pfad (lokal oder Azure)
            path = module_file.get("filename") or module_file.get("azure_params", {}).get("FILE_PATH") or "Nicht geladen"
            print(f"✅ Module: {path}")
            
        except Exception as e:
            errors.append(f"Module: {str(e)}")
            config["module"] = {"filename": None, "sheet": None, "azure_params": None, "columns": {}}
            print(f"❌ Module-Fehler: {e}")
        
        # === 3. Mapping ===
        try:
            mapping_file = globals().get("OTHER_FILES", {}).get("mapping", {})
            
            mapping_data = {
                "filename": mapping_file.get("filename"),
                "sheet": mapping_file.get("sheet"),
                "azure_params": mapping_file.get("azure_params"),
                "columns": config_manager.config.get("mapping", {}).get("columns", {})
            }
            config["mapping"] = mapping_data
            
            path = mapping_file.get("filename") or mapping_file.get("azure_params", {}).get("FILE_PATH") or "Nicht geladen"
            print(f"✅ Mapping: {path}")
            
        except Exception as e:
            errors.append(f"Mapping: {str(e)}")
            config["mapping"] = {"filename": None, "sheet": None, "azure_params": None, "columns": {}}
            print(f"❌ Mapping-Fehler: {e}")
        
        # === 4. Keywords ===
        try:
            keywords_file = globals().get("OTHER_FILES", {}).get("keywords", {})
            
            keywords_data = {
                "filename": keywords_file.get("filename"),
                "sheet": keywords_file.get("sheet"),
                "azure_params": keywords_file.get("azure_params"),
                "columns": {"Kategorie Suchvolumen": "Kategorie Suchvolumen"}
            }
            config["keywords"] = keywords_data
            
            path = keywords_file.get("filename") or keywords_file.get("azure_params", {}).get("FILE_PATH") or "Nicht geladen"
            print(f"✅ Keywords: {path}")
            
        except Exception as e:
            errors.append(f"Keywords: {str(e)}")
            config["keywords"] = {"filename": None, "sheet": None, "azure_params": None, "columns": {}}
            print(f"❌ Keywords-Fehler: {e}")
        
        # === 5. Systematik ===
        try:
            systematik_file = globals().get("OTHER_FILES", {}).get("systematik", {})
            
            systematik_data = {
                "filename": systematik_file.get("filename") or systematik_file.get("azure_params", {}).get("FILE_PATH"),
                "azure_params": systematik_file.get("azure_params"),
                "columns": {}
            }
            config["systematik"] = systematik_data
            
            path = systematik_data["filename"] or "Nicht geladen"
            print(f"✅ Systematik: {path}")
            
        except Exception as e:
            errors.append(f"Systematik: {str(e)}")
            config["systematik"] = {"filename": None, "azure_params": None, "columns": {}}
            print(f"❌ Systematik-Fehler: {e}")
        
        # === 6. Prompts ===
        try:
            prompts_file = globals().get("OTHER_FILES", {}).get("prompts", {})
            
            prompts_data = {
                "filename": prompts_file.get("filename") or prompts_file.get("azure_params", {}).get("FILE_PATH"),
                "azure_params": prompts_file.get("azure_params"),
                "columns": {}
            }
            config["prompts"] = prompts_data
            
            path = prompts_data["filename"] or "Nicht geladen"
            print(f"✅ Prompts: {path}")
            
        except Exception as e:
            errors.append(f"Prompts: {str(e)}")
            config["prompts"] = {"filename": None, "azure_params": None, "columns": {}}
            print(f"❌ Prompts-Fehler: {e}")
        
        # === ZUSAMMENFASSUNG ===
        print("\n" + "="*50)
        print("ZUSAMMENFASSUNG ZUM SPEICHERN:")
        print("="*50)
        
        summary_lines = []
        for key, data in config.items():
            path = data.get('filename') or (data.get('azure_params', {}).get('FILE_PATH') if data.get('azure_params') else None)
            if path:
                summary_lines.append(f"✅ {key}: {os.path.basename(str(path))}")
            else:
                summary_lines.append(f"❌ {key}: nicht geladen")
        
        summary = "\n".join(summary_lines)
        print(summary)
        
        if errors:
            print(f"\n⚠️ WARNUNGEN ({len(errors)}): {' | '.join(errors)}")
        
        # === SPEICHER-DIALOG ===
        fname, _ = QFileDialog.getSaveFileName(
            self, 
            "Config speichern", 
            CONFIG_FILE, 
            "JSON-Dateien (*.json);;Alle Dateien (*.*)"
        )
        if not fname:
            print("❌ Speichern abgebrochen durch Benutzer")
            return
        
        # Überschreiben-Abfrage
        if os.path.exists(fname):
            reply = QMessageBox.question(
                self, 
                "Überschreiben?", 
                f"Datei '{os.path.basename(fname)}' existiert bereits.\n\n"
                "Möchtest du sie überschreiben?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No  # Standard-Antwort
            )
            if reply == QMessageBox.No:
                print("❌ Speichern abgebrochen - Datei nicht überschrieben")
                return
        
        # === SPEICHERN ===
        try:
            with open(fname, "w", encoding="utf-8") as f:
                json.dump(config, f, ensure_ascii=False, indent=4)
            
            # Erfolgsmeldung mit Zusammenfassung
            msg_title = "Config erfolgreich gespeichert!"
            msg_body = f"Datei: {fname}\n\n{summary}"
            
            if errors:
                msg_body += f"\n\n⚠️ Hinweise: {len(errors)} Datei(en) hatten Probleme."
            
            QMessageBox.information(self, msg_title, msg_body)
            print(f"✅ Config erfolgreich gespeichert: {fname}")
            
            # Config-Manager neu laden
            config_manager.load()
            print("✅ Config-Manager neu geladen")
            
        except Exception as e:
            QMessageBox.critical(
                self, 
                "Speicher-Fehler", 
                f"Config konnte nicht gespeichert werden:\n\n{e}\n\n"
                "Mögliche Ursachen:\n"
                "• Keine Schreibrechte im Verzeichnis\n"
                "• Datei ist in einem anderen Programm geöffnet\n"
                "• Festplatte ist voll\n"
                "• Ungültiger Dateiname"
            )
            print(f"❌ Speichern fehlgeschlagen: {e}")
    
    def edit_column_names(self):
        """Dialog zum Bearbeiten der Spaltennamen ohne Excel-Datei"""
        # Zeige Dialog mit aktuellen Werten
        dlg = ColumnMappingSetupDialog(self)
        if dlg.exec() == QDialog.Accepted:
            # Config wird im Dialog bereits gespeichert
            # Aktualisiere globale Variablen
            initialize_column_config()
            QMessageBox.information(self, "Erfolg", "Spaltennamen wurden aktualisiert!")
    
    def map_columns_for(self, file_type: str):
        """Öffnet Spalten-Zuordnungs-Dialog für bereits geladene Datei"""
        df_map = {
            "modulkombi": globals().get("modulkombi_df"),
            "module": globals().get("module_df"),
            "mapping": globals().get("mapping_df"),
            "keywords": globals().get("keywords_df"),
            "systematik": None,  # JSON
            "prompts": None      # JSON
        }
        
        df = df_map.get(file_type)
        if df is None or (isinstance(df, pd.DataFrame) and df.empty):
            QMessageBox.warning(self, "Warnung", f"Bitte zuerst die {file_type}-Datei laden!")
            return
        
        dlg = ColumnMappingDialog(file_type, df, self)
        if dlg.exec() == QDialog.Accepted:
            mapping = dlg.get_mapping()
            # Speichere in Config
            config_manager.config[file_type]["columns"] = mapping
            config_manager.save()
            
            # Zeige kurze Bestätigung (statt Nachfrage)
            QMessageBox.information(
                self, 
                "Erfolg", 
                f"✅ Spalten-Zuordnung für '{file_type}' wurde gespeichert."
            )
    
    def load_modulkombi(self):
        """Variationen: modulkombi_df (Pflicht). Azure oder lokal."""
        global modulkombi_df, FILENAME, SHEETNAME, OTHER_FILES
        
        # Wenn Config vorhanden und Auto-Load aktiviert
        if config_manager.get("modulkombi.filename") and not self._loaded["modulkombi"]:
            reply = QMessageBox.question(self, "Auto-Load", 
                                       "Soll die Modulkombi-Datei aus der Config automatisch geladen werden?",
                                       QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                try:
                    df = pd.read_excel(config_manager.get("modulkombi.filename"),
                                     sheet_name=config_manager.get("modulkombi.sheet"))
                    FILENAME = config_manager.get("modulkombi.filename")
                    SHEETNAME = config_manager.get("modulkombi.sheet")
                    
                    # Spalten-Mapping anwenden
                    column_mapping = config_manager.get("modulkombi.columns", {})
                    if column_mapping:
                        df = df.rename(columns={v: k for k, v in column_mapping.items()})
                    
                    modulkombi_df = df
                    self._loaded["modulkombi"] = True
                    self._status_widgets["modulkombi"].setText("✅ Geladen (Config)")
                    QMessageBox.information(self, "Erfolg", "Modulkombi automatisch geladen!")
                    return
                except Exception as e:
                    QMessageBox.warning(self, "Auto-Load Fehlgeschlagen", f"Konnte nicht automatisch laden:\n{e}")
        
        while True:
            choice = self._ask_local_or_azure()
            if choice is None:
                return
            if choice == "Lokal":
                res = self._load_local_excel_with_sheet("Variationen (modulkombi_df)")
                if res is None:
                    continue
                df, fname, sheet = res
                ok, df = ensure_modulkombi_columns(self, df)  # Automatisches Mapping
                if not ok:
                    continue
                # SICHERHEITSCHECK: Gibt es die Dauer-Spalte?
                if DauerFeld not in df.columns:
                    QMessageBox.critical(self, "Fehler", f"Spalte '{DauerFeld}' wurde nicht gefunden!\n\n"
                                                      f"Gefundene Spalten: {', '.join(df.columns[:10])}...")
                    continue
                text_cols = set(REQUIRED_COLUMNS)
                existing = text_cols & set(df.columns)
                if existing:
                    df = df.astype({c: "string" for c in existing})
                modulkombi_df = df
                FILENAME = fname
                SHEETNAME = sheet
                OTHER_FILES["modulkombi"] = {"filename": fname, "sheet": sheet}
                self._loaded["modulkombi"] = True
                self._status_widgets["modulkombi"].setText("✅ Geladen")
                QMessageBox.information(self, "Erfolg", "Variationen erfolgreich geladen.")
                return
            else:  # Azure
                res = self._load_azure_file("Variationen (modulkombi_df)")
                if res is None:
                    continue
                data_bytes, params = res
                try:
                    xls = pd.ExcelFile(BytesIO(data_bytes))
                    sheet, ok = QInputDialog.getItem(self, "Blatt auswählen", "Bitte Blatt auswählen:", xls.sheet_names, 0, False)
                    if not ok:
                        continue
                    df = pd.read_excel(BytesIO(data_bytes), sheet_name=sheet)
                    ok_map, df = ensure_modulkombi_columns(self, df)
                    if not ok_map:
                        continue
                    text_cols = set(REQUIRED_COLUMNS)
                    existing = text_cols & set(df.columns)
                    if existing:
                        df = df.astype({c: "string" for c in existing})
                    modulkombi_df = df
                    FILENAME = f"azure:{params.get('FILE_PATH')}"
                    SHEETNAME = sheet
                    OTHER_FILES["modulkombi"] = {"azure_params": params, "sheet": sheet}
                    self._loaded["modulkombi"] = True
                    self._status_widgets["modulkombi"].setText("✅ Geladen (Azure)")
                    QMessageBox.information(self, "Erfolg", "Variationen (von Azure) erfolgreich geladen.")
                    return
                except Exception as e:
                    QMessageBox.critical(self, "Fehler beim Verarbeiten", str(e))
                    continue
                
        column_mapping = config_manager.get("modulkombi.columns", {})
        if column_mapping:
            df = df.rename(columns={v: k for k, v in column_mapping.items()})

    # ---------- Utilities to pick local or azure ----------
    def _ask_local_or_azure(self) -> Optional[str]:
        choice, ok = QInputDialog.getItem(self, "Quelle wählen", "Lade lokal oder via Azure?", ["Lokal"], 0, False)
                                         #["Lokal", "Azure"], 0, False)
        if not ok:
            return None
        return choice

    def _load_local_excel_with_sheet(self, title_hint="Excel-Datei") -> Optional[Tuple[pd.DataFrame, str, str]]:
        """Öffnet FileDialog, lässt Blatt wählen, liefert (df, filename, sheetname) oder None"""
        fname, _ = QFileDialog.getOpenFileName(self, f"{title_hint} wählen", "", "Excel Files (*.xlsx *.xls)")
        if not fname:
            return None
        try:
            xls = pd.ExcelFile(fname)
            sheet, ok = QInputDialog.getItem(self, "Blatt auswählen", "Bitte Blatt auswählen:", xls.sheet_names, 0, False)
            if not ok:
                return None
            df = pd.read_excel(fname, sheet_name=sheet)
            return df, fname, sheet
        except Exception as e:
            QMessageBox.critical(self, "Fehler", f"Fehler beim Lesen der Datei: {e}")
            return None

    def _load_azure_file(self, parent_label="Datei via Azure") -> Optional[Tuple[bytes, Dict[str, str]]]:
        """Fragt Azure-Parameter ab, lädt die Datei bytes und gibt (bytes, params) zurück."""
        dlg = AzureParamsDialog(self)
        params = dlg.get_params()
        if not params:
            return None
        try:
            data = download_file_from_sharepoint(params)
            return data, params
        except Exception as e:
            QMessageBox.critical(self, "Azure-Fehler", f"Fehler beim Laden von Azure/SharePoint:\n{e}")
            return None


    def load_module_df(self):
        """Kursinhalte (module_df) - Pflicht. Lokal oder Azure."""
        global module_df, OTHER_FILES
        while True:
            choice = self._ask_local_or_azure()
            if choice is None:
                return
            if choice == "Lokal":
                res = self._load_local_excel_with_sheet("Kursinhalte (module_df)")
                if res is None:
                    continue
                df, fname, sheet = res
                ok, df = ensure_module_columns(self, df)  # WICHTIG: Mapping anwenden
                if not ok:
                    continue
                module_df = df
                OTHER_FILES["module"] = {"filename": fname, "sheet": sheet}
                self._loaded["module"] = True
                self._status_widgets["module"].setText("✅ Geladen")
                QMessageBox.information(self, "Erfolg", "Kursinhalte erfolgreich geladen.")
                return
            else:
                res = self._load_azure_file("Kursinhalte (module_df)")
                if res is None:
                    continue
                data_bytes, params = res
                try:
                    xls = pd.ExcelFile(BytesIO(data_bytes))
                    sheet, ok = QInputDialog.getItem(self, "Blatt auswählen", "Bitte Blatt auswählen:", xls.sheet_names, 0, False)
                    if not ok:
                        continue
                    df = pd.read_excel(BytesIO(data_bytes), sheet_name=sheet)
                    ok_map, df = ensure_module_columns(self, df)  # WICHTIG: Mapping anwenden
                    if not ok_map:
                        continue
                    module_df = df
                    OTHER_FILES["module"] = {"azure_params": params, "sheet": sheet}
                    self._loaded["module"] = True
                    self._status_widgets["module"].setText("✅ Geladen (Azure)")
                    QMessageBox.information(self, "Erfolg", "Kursinhalte (Azure) erfolgreich geladen.")
                    return
                except Exception as e:
                    QMessageBox.critical(self, "Fehler beim Verarbeiten", str(e))
                    continue

    def load_mapping_df(self):
        """Mapping (mapping_df) - Pflicht. Lokal oder Azure."""
        global mapping_df, OTHER_FILES
        while True:
            choice = self._ask_local_or_azure()
            if choice is None:
                return
            if choice == "Lokal":
                res = self._load_local_excel_with_sheet("Mapping (mapping_df)")
                if res is None:
                    continue
                df, fname, sheet = res
                ok, df = ensure_mapping_columns(self, df)  # WICHTIG: Mapping anwenden
                if not ok:
                    continue
                mapping_df = df
                OTHER_FILES["mapping"] = {"filename": fname, "sheet": sheet}
                self._loaded["mapping"] = True
                self._status_widgets["mapping"].setText("✅ Geladen")
                QMessageBox.information(self, "Erfolg", "Mapping erfolgreich geladen.")
                return
            else:
                res = self._load_azure_file("Mapping (mapping_df)")
                if res is None:
                    continue
                data_bytes, params = res
                try:
                    xls = pd.ExcelFile(BytesIO(data_bytes))
                    sheet, ok = QInputDialog.getItem(self, "Blatt auswählen", "Bitte Blatt auswählen:", xls.sheet_names, 0, False)
                    if not ok:
                        continue
                    df = pd.read_excel(BytesIO(data_bytes), sheet_name=sheet)
                    ok_map, df = ensure_mapping_columns(self, df)  # WICHTIG: Mapping anwenden
                    if not ok_map:
                        continue
                    mapping_df = df
                    OTHER_FILES["mapping"] = {"azure_params": params, "sheet": sheet}
                    self._loaded["mapping"] = True
                    self._status_widgets["mapping"].setText("✅ Geladen (Azure)")
                    QMessageBox.information(self, "Erfolg", "Mapping (Azure) erfolgreich geladen.")
                    return
                except Exception as e:
                    QMessageBox.critical(self, "Fehler beim Verarbeiten", str(e))
                    continue

    def load_keywords_df(self):
        """Keywords (keywords_df) - Pflicht. Lokal oder Azure. Nachladen volume_score."""
        global keywords_df, volume_map, OTHER_FILES
        while True:
            choice = self._ask_local_or_azure()
            if choice is None:
                return
            if choice == "Lokal":
                res = self._load_local_excel_with_sheet("Keywords (keywords_df)")
                if res is None:
                    continue
                df, fname, sheet = res
                keywords_df = df
                # compute volume_map
                volume_map = {'High': 3, 'Medium': 2, 'Low': 1}
                try:
                    if 'Kategorie Suchvolumen' in keywords_df.columns:
                        keywords_df['volume_score'] = keywords_df['Kategorie Suchvolumen'].map(volume_map)
                    else:
                        QMessageBox.warning(self, "Warnung", "Spalte 'Kategorie Suchvolumen' nicht gefunden in keywords; 'volume_score' nicht berechnet.")
                    OTHER_FILES["keywords"] = {"filename": fname, "sheet": sheet}
                    self._loaded["keywords"] = True
                    self._status_widgets["keywords"].setText("✅ Geladen")
                    QMessageBox.information(self, "Erfolg", "Keywords erfolgreich geladen.")
                    return
                except Exception as e:
                    QMessageBox.critical(self, "Fehler beim Verarbeiten", str(e))
                    continue
            else:
                res = self._load_azure_file("Keywords (keywords_df)")
                if res is None:
                    continue
                data_bytes, params = res
                try:
                    xls = pd.ExcelFile(BytesIO(data_bytes))
                    sheet, ok = QInputDialog.getItem(self, "Blatt auswählen", "Bitte Blatt auswählen:", xls.sheet_names, 0, False)
                    if not ok:
                        continue
                    df = pd.read_excel(BytesIO(data_bytes), sheet_name=sheet)
                    keywords_df = df
                    volume_map = {'High': 3, 'Medium': 2, 'Low': 1}
                    if 'Kategorie Suchvolumen' in keywords_df.columns:
                        keywords_df['volume_score'] = keywords_df['Kategorie Suchvolumen'].map(volume_map)
                    else:
                        QMessageBox.warning(self, "Warnung", "Spalte 'Kategorie Suchvolumen' nicht gefunden in keywords; 'volume_score' nicht berechnet.")
                    OTHER_FILES["keywords"] = {"azure_params": params, "sheet": sheet}
                    self._loaded["keywords"] = True
                    self._status_widgets["keywords"].setText("✅ Geladen")
                    QMessageBox.information(self, "Erfolg", "Keywords (Azure) erfolgreich geladen.")
                    return
                except Exception as e:
                    QMessageBox.critical(self, "Fehler beim Verarbeiten", str(e))
                    continue

    def load_systematik(self):
        """systematiken.json - Pflicht - lokal oder Azure"""
        global SYSTEMATIK, OTHER_FILES
        while True:
            choice = self._ask_local_or_azure()
            if choice is None:
                return
            if choice == "Lokal":
                fname, _ = QFileDialog.getOpenFileName(self, "systematiken.json wählen", "", "JSON Files (*.json)")
                if not fname:
                    continue
                try:
                    text = Path(fname).read_text(encoding="utf-8")
                    SYSTEMATIK = json.loads(text)
                    OTHER_FILES["systematik"] = {"filename": fname}
                    self._loaded["systematik"] = True
                    self._status_widgets["systematik"].setText("✅ Geladen")
                    QMessageBox.information(self, "Erfolg", "systematiken.json erfolgreich geladen.")
                    return
                except Exception as e:
                    QMessageBox.critical(self, "Fehler beim Lesen JSON", str(e))
                    continue
            else:
                res = self._load_azure_file("systematiken.json")
                if res is None:
                    continue
                data_bytes, params = res
                try:
                    text = data_bytes.decode("utf-8")
                    SYSTEMATIK = json.loads(text)
                    OTHER_FILES["systematik"] = {"azure_params": params}
                    self._loaded["systematik"] = True
                    self._status_widgets["systematik"].setText("✅ Geladen")
                    QMessageBox.information(self, "Erfolg", "systematiken.json (Azure) erfolgreich geladen.")
                    return
                except Exception as e:
                    QMessageBox.critical(self, "Fehler beim Verarbeiten JSON", str(e))
                    continue
                
    def load_prompts(self):
        """prompts.json - Pflicht - lokal oder Azure"""
        global prompts, OTHER_FILES  # WICHTIG: prompts statt SYSTEMATIK!
        
        while True:
            choice = self._ask_local_or_azure()
            if choice is None:
                return
            
            if choice == "Lokal":
                fname, _ = QFileDialog.getOpenFileName(self, "prompts.json wählen", "", "JSON Files (*.json)")
                if not fname:
                    continue
                try:
                    text = Path(fname).read_text(encoding="utf-8")
                    prompts = json.loads(text)  # WICHTIG: prompts statt SYSTEMATIK!
                    OTHER_FILES["prompts"] = {"filename": fname}
                    self._loaded["prompts"] = True
                    self._status_widgets["prompts"].setText("✅ Geladen")
                    QMessageBox.information(self, "Erfolg", "prompts.json erfolgreich geladen.")
                    return
                except Exception as e:
                    QMessageBox.critical(self, "Fehler beim Lesen JSON", str(e))
                    continue
            
            else:  # Azure
                res = self._load_azure_file("prompts.json")
                if res is None:
                    continue
                data_bytes, params = res
                try:
                    text = data_bytes.decode("utf-8")
                    prompts = json.loads(text)
                    OTHER_FILES["prompts"] = {"azure_params": params}
                    self._loaded["prompts"] = True
                    self._status_widgets["prompts"].setText("✅ Geladen")
                    QMessageBox.information(self, "Erfolg", "prompts.json (Azure) erfolgreich geladen.")
                    return
                except Exception as e:
                    QMessageBox.critical(self, "Fehler beim Verarbeiten JSON", str(e))
                    continue

    # ---------- Abschluss / Validierung ----------
    def on_ok(self):
        # Prüfe, ob alles geladen ist
        missing = [k for k, v in self._loaded.items() if not v]
        if missing:
            QMessageBox.critical(self, "Nicht vollständig", f"Folgende Dateien fehlen noch: {missing}. Laden abbrechen oder Dateien vollständig laden.")
            return
        # Alle Dateien geladen -> accept und raus
        self.accept()


def create_termine(options, FILENAME: str, SHEETNAME: str, Termine_Feld: str, debug_level: int = 0):
    print("⏳ Erstelle Termine...")
    """
    Debug-Variante von create_termine:
      - debug_level: 0 (nur Fehler), 1 (Basis-Logs), 2 (sehr ausführlich)
      - nutzt Termine_Feld zum Finden der Spalte (kein harter String)
      - schreibt in eine temporäre Datei, saved und reopened -> vergleicht Werte
    """

    log("⏳ Erstelle Termine...")
    try:
        # -------------------------
        # 1) Termine aus Status Select laden
        # -------------------------
        termine_df = pd.read_excel(
            FILENAME,
            sheet_name="Status Select",
            header=0,
            usecols=["Dauer", "Termine"]
        )
        if debug_level == 2:
            log(f"[Status Select] shape: {termine_df.shape}")
            log(f"[Status Select] dtypes:\n{termine_df.dtypes}")
            log(f"[Status Select] head:\n{termine_df.head()}")

        termine_df["Dauer"] = pd.to_numeric(termine_df["Dauer"], errors="coerce").round().astype("Int64")
        termine_df = termine_df.dropna(subset=["Dauer"])
        if debug_level == 2:
            log(f"[Status Select] after cleaning shape: {termine_df.shape}", 2)
            log(f"[Status Select] available durations: {sorted(termine_df['Dauer'].unique().tolist())}", 2)

        # Lookup arrays
        avail_durs = termine_df["Dauer"].sort_values().to_numpy()
        term_strings = termine_df.set_index("Dauer")["Termine"]
        if debug_level == 2:
            log(f"[Lookup] avail_durs: {avail_durs}", 2)
            log(f"[Lookup] sample term_strings:\n{term_strings.head(10)}", 2)

        def closest_term(dauer):
            """Liefert den exakten Zellinhalt zur nächsten verfügbaren Dauer."""
            if pd.isna(dauer):
                return np.nan
            try:
                # sicherstellen, dass dauer numerisch ist
                d = float(dauer)
            except Exception:
                return np.nan
            idx = np.searchsorted(avail_durs, d)
            if idx == 0:
                closest = avail_durs[0]
            elif idx == len(avail_durs):
                closest = avail_durs[-1]
            else:
                before, after = avail_durs[idx - 1], avail_durs[idx]
                closest = before if abs(d - before) <= abs(after - d) else after
            # term_strings ist Series mit Index = Dauer (Int64); Zugriff absichern
            try:
                return term_strings.loc[closest]
            except KeyError:
                # Fallback: nearest by index
                return term_strings.iloc[(np.abs(termine_df["Dauer"] - closest)).argmin()]

        # -------------------------
        # 2) Variationen laden und Termine zuordnen
        # -------------------------
        modulkombi_df = pd.read_excel(FILENAME, sheet_name=SHEETNAME, header=0)
        if debug_level == 2:
            log(f"[{SHEETNAME}] shape: {modulkombi_df.shape}", 2)
            log(f"[{SHEETNAME}] columns: {list(modulkombi_df.columns)}", 2)
            log(f"[{SHEETNAME}] dtypes:\n{modulkombi_df.dtypes}", 2)

        # safe cast Dauer in Tagen
        if "Dauer in Tagen" not in modulkombi_df.columns:
            raise KeyError("Spalte 'Dauer in Tagen' nicht in Sheet gefunden.")
        modulkombi_df["Dauer in Tagen"] = pd.to_numeric(modulkombi_df["Dauer in Tagen"], errors="coerce").round().astype("Int64")
        if debug_level == 2:
            log(f"[{SHEETNAME}] 'Dauer in Tagen' stats:\n{modulkombi_df['Dauer in Tagen'].describe(include='all')}", 2)

        # WICHTIG: in neue Spalte schreiben, nicht in 'Dauer in Tagen' überschreiben
        modulkombi_df["ESTHER_Termine_DEBUG"] = modulkombi_df["Dauer in Tagen"].apply(closest_term)
        if debug_level == 2:
            log(f"[{SHEETNAME}] ESTHER_Termine_DEBUG head:\n{modulkombi_df[['Dauer in Tagen','ESTHER_Termine_DEBUG']].head(15)}", 2)

        # -------------------------
        # 3) Excel öffnen + Spalte finden (Termine_Feld verwenden)
        # -------------------------
        wb = openpyxl.load_workbook(FILENAME)
        if SHEETNAME not in wb.sheetnames:
            raise KeyError(f"Sheet '{SHEETNAME}' nicht in Workbook.")
        ws = wb[SHEETNAME]

        # Header-Zeile lesen
        hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if debug_level == 2:
            log(f"[Excel] Header: {hdr}", 2)

        # Termine_Feld verwenden (user input) — robustes Matching
        # Versuche exakten Match, dann strip Match (whitespace) dann case-insensitive
        term_col_idx = None
        try:
            term_col_idx = hdr.index(Termine_Feld) + 1
        except ValueError:
            # try stripped
            stripped_hdr = [h.strip() if isinstance(h, str) else h for h in hdr]
            try:
                term_col_idx = stripped_hdr.index(Termine_Feld.strip()) + 1
            except ValueError:
                # case-insensitive search
                lowered = [h.lower() if isinstance(h, str) else h for h in hdr]
                try:
                    term_col_idx = lowered.index(Termine_Feld.lower()) + 1
                except ValueError:
                    raise ValueError(f"Spalte '{Termine_Feld}' nicht gefunden im Header: {hdr}")

        if debug_level == 2:
            log(f"[Excel] Termine-Spalte index (1-based): {term_col_idx}", 1)

        # -------------------------
        # 4) Werte in Zellen schreiben (erst nur in-memory und log)
        # -------------------------
        # Prepare values to write
        write_values = []
        for idx, val in enumerate(modulkombi_df["ESTHER_Termine_DEBUG"], start=2):
            if pd.isna(val):
                write_values.append("")
            else:
                # sicherstellen, dass Zeilenumbrüche wirklich \n enthalten
                v = str(val)
                write_values.append(v)

        if debug_level == 2:
            log(f"[Write] Anzahl zu schreibender Zeilen: {len(write_values)} (erste 10): {write_values[:10]}", 2)

        # Schreibe in Workbook
        for row_idx, value in enumerate(write_values, start=2):
            cell = ws.cell(row=row_idx, column=term_col_idx, value=value)
            cell.alignment = Alignment(wrapText=True, vertical="top")

        # -------------------------
        # 5) Save in temporäre Datei und überprüfe
        # -------------------------
        dirn = os.path.dirname(os.path.abspath(FILENAME)) or "."
        temp_fd, temp_path = tempfile.mkstemp(prefix="debug_termine_", suffix=".xlsx", dir=dirn)
        os.close(temp_fd)  # nur Pfad benötigt
        if debug_level == 2:
            log(f"[Save] temporäre Datei: {temp_path}", 1)

        try:
            wb.save(temp_path)
        except PermissionError as pe:
            raise PermissionError(f"Fehler beim Speichern: Datei möglicherweise von Excel geöffnet. ({pe})")

        # Re-open to verify
        wb2 = openpyxl.load_workbook(temp_path, data_only=True)
        ws2 = wb2[SHEETNAME]
        # read back the written column for the same rows and compare sample
        read_back = []
        for row_idx in range(2, 2 + len(write_values)):
            read_back.append(ws2.cell(row=row_idx, column=term_col_idx).value or "")

        # Vergleich
        mismatches = []
        for i, (expected, actual) in enumerate(zip(write_values, read_back)):
            if str(expected) != str(actual):
                mismatches.append((i+2, expected, actual))  # store row num (+2 offset)

        if mismatches:
            log(f"⚠️ MISMATCHES gefunden: {len(mismatches)} Zeilen unterscheiden sich (zeige bis 20):", 1)
            for m in mismatches[:20]:
                log(f"  Row {m[0]}: expected={repr(m[1])}  actual={repr(m[2])}", 1)
        else:
            log("✅ Nach Save: alle geprüften Werte stimmen überein (tempfile).", 1)

        # -------------------------
        # 6) optional: überschreibe Original (nur wenn Optionen das erlauben)
        # -------------------------
        if getattr(options, "overwrite_original", True):

            # move temp over original
            try:
                shutil.move(temp_path, FILENAME)
                log(f"✅ Temporäre Datei erfolgreich nach {FILENAME} verschoben.", 1)
            except Exception as e:
                log(f"❌ Fehler beim Überschreiben der Originaldatei: {e}", 0)
                # falls move fehlschlägt, entferne temp
                try:
                    os.remove(temp_path)
                except Exception:
                    pass
                raise
        else:
            log(f"[Info] overwrite_original=False -> temp Datei bleibt unter: {temp_path}", 1)

    except Exception as e:
        # umfangreiche Fehlermeldung + Traceback
        print(f"🚨 Fehler in create_termine: {e}")
        traceback.print_exc()
        raise

    log("⏳ Termine erfolgreich erstellt")






def search_mapping_any_form(base_id: str) -> pd.DataFrame:
    """
    Return rows from mapping_df that contain any plausible form of base_id
    in any of the mapping columns.
    base_id may be like 'A63401' or 'A634-01' etc.
    """
    # Build normalized literal variants
    variants = extract_relevant_id(base_id, all_variants=True)
    # boolean mask with same index as mapping_df
    mask = pd.Series(False, index=mapping_df.index)

    # search each column for each variant using literal contains (regex=False)
    for col in ['Modulnummer Alt', 'Modulnummer Neu', 'Modulnummer HP']:
        ser = mapping_df[col].astype(str).str.upper()
        for v in variants:
            # use regex=False to treat v as literal (safe for '-' and '/')
            mask |= ser.str.contains(v, na=False, regex=False)

    return mapping_df[mask]


# --- main finder -------------------------------------------------------
def sanitize_id(raw: str) -> str:
    """Bereinigt ID: Entfernt Leerzeichen um Trenner und nach Präfixen"""
    s = str(raw).strip().upper()
    # Leerzeichen um / - _ entfernen
    s = re.sub(r'\s*([/_-])\s*', r'\1', s)
    # Leerzeichen nach D_ oder KN- entfernen
    s = re.sub(r'(D_|KN-)\s+', r'\1', s)
    # Mehrfache Unterstriche bereinigen
    s = re.sub(r'_+', r'_', s)
    return s

def extract_relevant_id(raw: str, *, all_variants: bool = False):
    """
    Erzeugt deterministische Varianten:
    - Mit/ohne Trenner (- /)
    - Mit/ohne führendes A/V
    - Mit alternativem Prefix (A ↔ V)
    - Mit/ohne Präfixe (D_, KN-, KN-D_)
    """
    s = sanitize_id(raw)
    
    m = re.search(r'^([AV]?)(\d{3,4})[-/]?(\d{1,4})$', s, re.I)
    if not m:
        return [s] if all_variants else s
    
    prefix, num, tail = m.groups()
    prefix = prefix or ''
    
    base_variants = [
        f"{prefix}{num}{tail}",
        f"{prefix}{num}-{tail}",
        f"{prefix}{num}/{tail}"
    ]
    
    if not all_variants:
        return base_variants[0]
    
    all_variants_set = set(base_variants)
    
    # Varianten ohne führendes A/V
    if prefix in ['A', 'V']:
        no_prefix = [f"{num}{tail}", f"{num}-{tail}", f"{num}/{tail}"]
        all_variants_set.update(no_prefix)
        
        # Alternativer Prefix
        alt_prefix = 'V' if prefix == 'A' else 'A'
        alt_variants = [
            f"{alt_prefix}{num}{tail}",
            f"{alt_prefix}{num}-{tail}",
            f"{alt_prefix}{num}/{tail}"
        ]
        all_variants_set.update(alt_variants)
    
    return list(all_variants_set)

def extract_core_id(s: str):
    # entfernt alle Präfixe wie D_, A_, KN-D_, KN_D-, usw.
    s = re.sub(r'^[A-Z]{1,3}[-_]', '', s)   # entfernt z.B. D_, KN-, A-, …
    return s

def get_all_search_variants(raw_id: str) -> list:
    """
    Erzeugt ALLE Suchvarianten inkl. aller Präfix-Kombinationen.
    Deckt auch Fälle wie D_A805/68, KN-D_A805/68, D_805/68, etc. ab.
    """
    core_variants = extract_relevant_id(raw_id, all_variants=True)
    
    # Alle möglichen Präfix-Kombinationen
    prefixes = ['', 'D_', 'KN-', 'KN-D_']
    
    # Sammle alle Kombinationen
    all_variants = set()
    for core in core_variants:
        # Wenn core bereits Präfix enthält, nimm es nur einmal
        if any(core.startswith(p.rstrip('-_')) for p in prefixes):
            all_variants.add(core)
        else:
            for prefix in prefixes:
                all_variants.add(prefix + core)
    
    # Sortiere nach Wahrscheinlichkeit (KN-D_ > D_ > KN- > ohne)
    def sort_key(v):
        score = 0
        if v.startswith('KN-D_'):
            score = 30
        elif v.startswith('D_'):
            score = 20
        elif v.startswith('KN-'):
            score = 10
        if any(c.isalpha() for c in v.replace('_', '').replace('-', '')):
            score += 5  # Mit Buchstaben bevorzugen
        return -score
    
    return sorted(all_variants, key=sort_key)


def find_module_rows(worker, module_ids, kurs_title="Unbekannt", idx=-1, debug=False) -> pd.DataFrame:
    """
    Fuzzy-first mit progressiver Filterung:
    1. Fuzzy contains (schnell, tolerant)
    2. Falls >1 Treffer → exakte Filterung
    3. Falls >1 Treffer → Wortgrenzen-Filter
    4. Falls keine Treffer → Mapping mit allen Varianten
    """
    result_records = []
    seen_keys = set()
    unfound_titles = []

    def add_records(matches):
        """Fügt matches hinzu, falls noch nicht gesehen"""
        nonlocal result_records, seen_keys
        if not matches.empty:
            for rec in matches.to_dict('records'):
                key = f"{rec.get('Modulnummer1')}_{rec.get('Modulnummer2')}_{rec.get('Titel')}"
                if key not in seen_keys:
                    result_records.append(rec)
                    seen_keys.add(key)
            return True
        return False

    for i, raw_id in enumerate(module_ids, start=1):
        original_mod_id = str(raw_id).strip()
        variants = get_all_search_variants(raw_id)
        
        if debug:
            print(f"\n--- processing {i}: {original_mod_id}  → {len(variants)} variants")
            if len(variants) <= 12:
                print(f"  variants: {variants}")

        hit = False
        potential_matches = []
        
        # === STUFE 1: FUZZY contains-Suche ===
        if debug:
            print(f"  🎯 Stage 1: Fuzzy search")
        
        for mod_id in variants:
            # Suche in beiden Spalten (case-insensitive, ohne regex)
            for col in ['Modulnummer1', 'Modulnummer2']:
                # Entferne Leerzeichen für robustere Suche
                series = module_df[col].astype(str).apply(sanitize_id)
                matches = module_df[series.str.contains(mod_id, case=False, na=False, regex=False)]
                if not matches.empty:
                    potential_matches.append(matches)
        
        if potential_matches:
            combined = pd.concat(potential_matches).drop_duplicates()
            
            if debug:
                print(f"    📊 {len(combined)} potential matches")
            
            # Genau 1 Treffer? Perfekt!
            if len(combined) == 1:
                if debug:
                    print(f"    ✅ Single fuzzy match")
                hit = add_records(combined)
            else:
                # Mehrere Treffer → Filter
                if debug:
                    print(f"    ⚠ Applying exact filter...")
                
                # === STUFE 2: Exakte Filterung ===
                filtered_hits = []
                for mod_id in variants:
                    mod_id_upper = mod_id.upper()
                    
                    # Exaktes Match in Modulnummer1
                    exact1 = combined[combined['Modulnummer1'].astype(str).apply(sanitize_id) == mod_id_upper]
                    if not exact1.empty:
                        filtered_hits.append(exact1)
                    
                    # Exaktes Match in Modulnummer2
                    exact2 = combined[combined['Modulnummer2'].astype(str).apply(sanitize_id) == mod_id_upper]
                    if not exact2.empty:
                        filtered_hits.append(exact2)
                
                if filtered_hits:
                    combined_exact = pd.concat(filtered_hits).drop_duplicates()
                    
                    if len(combined_exact) == 1:
                        if debug:
                            print(f"    ✅ Single exact match")
                        hit = add_records(combined_exact)
                    else:
                        # === STUFE 3: Wortgrenzen-Filter ===
                        if debug:
                            print(f"    ⚠ {len(combined_exact)} exact matches, word boundary filter...")
                        
                        for mod_id in variants:
                            pattern = r'\b' + re.escape(mod_id) + r'\b'
                            
                            for col in ['Modulnummer1', 'Modulnummer2']:
                                series = combined_exact[col].astype(str).apply(sanitize_id)
                                word_matches = combined_exact[series.str.contains(pattern, na=False, regex=True)]
                                
                                if not word_matches.empty:
                                    hit = add_records(word_matches)
                                    if hit:
                                        if debug:
                                            print(f"    ✅ Word boundary match in {col}")
                                        break
                            if hit:
                                break
                else:
                    if debug:
                        print(f"    ⚠ No exact matches, trying word boundary on all...")
                    # Fallback: Wortgrenzen auf alle fuzzy Treffer
                    for mod_id in variants:
                        pattern = r'\b' + re.escape(mod_id) + r'\b'
                        
                        for col in ['Modulnummer1', 'Modulnummer2']:
                            series = combined[col].astype(str).apply(sanitize_id)
                            word_matches = combined[series.str.contains(pattern, na=False, regex=True)]
                            
                            if not word_matches.empty:
                                hit = add_records(word_matches)
                                if debug:
                                    print(f"    ✅ Word boundary match in {col}")
                                break
                        if hit:
                            break
        
        # === STUFE 4: MAPPING ===
        if not hit:
            if debug:
                print(f"  🎯 Stage 4: Mapping lookup")
            
            mapping_rows = pd.DataFrame()
            for v in variants:
                found = search_mapping_any_form(v)
                if not found.empty:
                    mapping_rows = pd.concat([mapping_rows, found]).drop_duplicates()
            
            if not mapping_rows.empty:
                if debug:
                    print(f"    🔁 {len(mapping_rows)} mapping rows found")
                
                target_ids = set()
                for _, row in mapping_rows.iterrows():
                    # Prüfe Quell-Spalten
                    source_match = False
                    for scol in ['Modulnummer Alt', 'Modulnummer HP']:
                        cell = str(row.get(scol, '')).strip().upper()
                        # Prüfe exakte Übereinstimmung mit unseren Varianten
                        cell_vals = [sanitize_id(c.strip()) for c in cell.split(",")]
                        for v in variants:
                            v_clean = sanitize_id(v)
                            if any(v_clean in c or c in v_clean for c in cell_vals):
                                source_match = True
                                break

                    
                    if source_match:
                        # Ziel-IDs aus "Modulnummer Neu" (komma-getrennt)
                        neu_cell = str(row.get('Modulnummer Neu', '')).strip()
                        if neu_cell:
                            target_ids.update([t.strip() for t in neu_cell.split(',')])
                
                # Suche jede Ziel-ID
                if debug:
                    print("Target IDs: ")
                    print(target_ids)
                for target_id in target_ids:
                    core = extract_core_id(target_id)
                    target_variants = get_all_search_variants(core)
                    if debug:
                        print(f"      searching target: {target_id} → {len(target_variants)} variants")
                    
                    for tvar in target_variants:
                        # Wiederholte Stufen 1-3 für Ziel-ID
                        for col in ['Modulnummer1', 'Modulnummer2']:
                            series = module_df[col].astype(str).apply(sanitize_id)
                            matches = module_df[series.str.contains(tvar, case=False, na=False, regex=False)]
                            
                            if len(matches) == 1:
                                hit = add_records(matches)
                                if debug:
                                    print(f"        ✅ single fuzzy match for {tvar}")
                                break
                            elif len(matches) > 1:
                                # Exakte Filterung
                                filtered = matches[series == tvar.upper()]
                                if not filtered.empty:
                                    hit = add_records(filtered)
                                    if debug:
                                        print(f"        ✅ exact match for {tvar}")
                                    break
                                # Wortgrenzen-Filter
                                pattern = r'\b' + re.escape(tvar) + r'\b'
                                filtered = matches[series.str.contains(pattern, na=False, regex=True)]
                                if not filtered.empty:
                                    hit = add_records(filtered)
                                    if debug:
                                        print(f"        ✅ word boundary match for {tvar}")
                                    break
                        if hit:
                            break
                    if hit:
                        break
        
        if not hit:
            unfound_titles.append((i, variants))
            if debug:
                print(f"  ⚠ No match after all stages")

    # AFTER the loop: handle all unfound titles in a single dialog
    if unfound_titles:
        # Nur EINE Variante pro Eintrag – bevorzugt mit '/'
        prefill = []
        for idx_row, variants in unfound_titles:
            if isinstance(variants, str):
                variants = [variants]
            # Wähle nur eine Variante (bevorzugt die mit '/')
            chosen = next((v for v in variants if "/" in v), variants[0])
            prefill.append((idx_row, [chosen]))
    
        worker.ask_module_correction.emit(prefill)
        corrected_map = worker.wait_for_gui_result()
    
        if corrected_map is None:
            raise UserCancelledError("Benutzer hat die Korrektur abgebrochen")
    
        # Korrigierte Einträge erneut prüfen
        for idx_row, corrected_list in corrected_map.items():
            if corrected_list is None:
                raise UserCancelledError(f"Zeile {idx_row} übersprungen")
            for new_mod in corrected_list:
                matched = match_single_id(new_mod, result_records, seen_keys, debug=debug)
                if not matched:
                    if debug:
                        print(f"  ⚠ Nach Korrektur weiterhin kein Treffer für {new_mod} (Zeile {idx_row})")
                    raise ModuleNotFoundError(f"  ⚠ Nach Korrektur weiterhin kein Treffer für {new_mod} (Zeile {idx_row})")
                    



    # final assembly
    if result_records:
        return pd.DataFrame(result_records).reset_index(drop=True)
    return pd.DataFrame()


def match_single_id(mod_id: str, result_records: list, seen_keys: set, debug: bool = False) -> bool:
    """
    Try to find mod_id in module_df (Modulnummer1 / Modulnummer2) or via mapping.
    If matches found, append records (dicts) to result_records and add keys to seen_keys.
    Returns True if any match found (hit), False otherwise.
    """

    mod_id = str(mod_id).strip().upper()
    # direct search Modulnummer1
    ser1 = module_df['Modulnummer1'].astype(str).str.upper()
    match1 = module_df[ser1.str.contains(mod_id, na=False, regex=False)]
    if not match1.empty:
        for rec in match1.to_dict('records'):
            key = f"{rec.get('Modulnummer1')}_{rec.get('Modulnummer2')}_{rec.get('Titel')}"
            if key not in seen_keys:
                result_records.append(rec)
                seen_keys.add(key)
        if debug:
            print(f"    ✅ direct found in Modulnummer1: {mod_id} ({len(match1)} rows)")
        return True

    # direct search Modulnummer2
    ser2 = module_df['Modulnummer2'].astype(str).str.upper()
    match2 = module_df[ser2.str.contains(mod_id, na=False, regex=False)]
    if not match2.empty:
        for rec in match2.to_dict('records'):
            key = f"{rec.get('Modulnummer1')}_{rec.get('Modulnummer2')}_{rec.get('Titel')}"
            if key not in seen_keys:
                result_records.append(rec)
                seen_keys.add(key)
        if debug:
            print(f"    ✅ direct found in Modulnummer2: {mod_id} ({len(match2)} rows)")
        return True

    # mapping lookup
    mapping_rows = search_mapping_any_form(mod_id)
    if mapping_rows.empty:
        if debug:
            print(f"    🔍 mapping lookup empty for {mod_id}")
        return False

    # follow mapping rows to module_df
    if debug:
        print(f"    🔁 mapping rows found ({len(mapping_rows)}) for {mod_id}")

    all_related_original = set()
    for _, mapping_row in mapping_rows.iterrows():
        for col in ['Modulnummer Alt', 'Modulnummer Neu', 'Modulnummer HP']:
            val = mapping_row.get(col)
            if pd.notna(val):
                all_related_original.add(str(val).strip().upper())

    for original_mapping_id in sorted(all_related_original):
        mapping_variants = [v.strip().upper() for v in extract_relevant_id(original_mapping_id, all_variants=True)]
        if debug:
            print(f"      mapping id {original_mapping_id} -> variants {mapping_variants}")
        for new_id_clean in mapping_variants:
            ser1 = module_df['Modulnummer1'].astype(str).str.upper()
            ser2 = module_df['Modulnummer2'].astype(str).str.upper()
            submatch1 = module_df[ser1.str.contains(new_id_clean, na=False, regex=False)]
            submatch2 = module_df[ser2.str.contains(new_id_clean, na=False, regex=False)]
            for sub in (submatch1, submatch2):
                if not sub.empty:
                    for rec in sub.to_dict('records'):
                        key = f"{rec.get('Modulnummer1')}_{rec.get('Modulnummer2')}_{rec.get('Titel')}"
                        if key not in seen_keys:
                            result_records.append(rec)
                            seen_keys.add(key)
                    if debug:
                        print(f"        ✅ found via mapping -> {original_mapping_id} (variant {new_id_clean})")
                    return True
    return False




def strip_intro(text):
    """
    Remove intro text from a block and start from the first relevant marker like * or ! or double newline.
    """
    if not text:
        return ""

    # Try to find the start of real content
    pattern = r"(^.*?\n)(?=(!|\*|\n))"  # everything up to first ! or * or \n\n
    match = re.search(pattern, text, flags=re.DOTALL)
    if match:
        return text[match.end():].lstrip()
    return text

def html_to_clean_text(html_content):
    if pd.isna(html_content):
        return ""
    html_content = html.unescape(html_content)
    soup = BeautifulSoup(html_content, "html.parser")
    return soup.get_text(separator="\n").strip()

def clean_zielgruppe_inputs(zielgruppen_raw: list[str]) -> list[str]:
    """
    Entfernt Zeilen oder Satzteile, die alternative Kurse erwähnen,
    bevor sie in die AI-Zusammenfassung gehen.
    """
    cleaned_blocks = []
    
    for block in zielgruppen_raw:
        lines = block.splitlines()
        cleaned_lines = []
        
        for line in lines:
            # Zeile komplett entfernen, wenn sie eindeutig einen Kurs erwähnt
            if re.search(r"(alternativ\s*:?\s*Kurs|Besuch des Kurses|Kurs [A-Z])", line, re.IGNORECASE):
                continue
            # Auch Inhalte in Klammern wie "(alternativ: Kurs ...)" entfernen
            line = re.sub(r"\(.*?alternativ.*?kurs.*?\)", "", line, flags=re.IGNORECASE)
            cleaned_lines.append(line.strip())
        
        cleaned_blocks.append("\n".join(cleaned_lines).strip())
    
    return cleaned_blocks


# =========== Keywords aus Titel rausfinden
def normalize_text(s: str) -> str:
    if not isinstance(s, str):
        raise TypeError(f"normalize_text() erwartet str, bekam {type(s)}: {s!r}")
    return s.lower().replace("...", " ")
    s = s.lower()
    s = re.sub(r"[/:\-]", " ", s)  # replace separators with spaces
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^a-z0-9äöüß ]", "", s.lower().strip).strip()
    return s

# ------------------------------------------------------------------
#    Normalisierung inkl. Titel-Grammatik (erstes + letztes Wort groß)
# ------------------------------------------------------------------
def _protect_composites(text: str) -> tuple[str, list[tuple[str, str]]]:
    """
    Ersetzt alle COMPOSITE_EXCEPTIONS durch spaCy-sichere Platzhalter,
    damit sie später exakt wieder eingesetzt werden können.
    """
    replacements = []
    protected_parts = []
    last_end = 0

    for match in _COMPOSITES_RE.finditer(text):
        start, end = match.span()
        placeholder = f"COMPPLACEHOLDER{len(replacements)}"
        protected_parts.append(text[last_end:start])
        protected_parts.append(placeholder)
        replacements.append((placeholder, match.group(0)))
        last_end = end

    protected_parts.append(text[last_end:])
    protected = "".join(protected_parts)

    return protected, replacements

# ------------------------------------------------------------------
#    Normalisierung mit SpaCy
# ------------------------------------------------------------------
def normalize_with_spacy(title: str) -> str:
    """
    Korrigiert die Groß-/Kleinschreibung deutscher Titel unter Beachtung von:
      - zusammengesetzten Fachbegriffen (COMPOSITE_EXCEPTIONS)
      - SPECIAL_WORDS (z. B. KI → KI, SAP → SAP)
      - SHORT_WORD_WHITELIST (z. B. SAP, AWS …)
      - Stopwörtern (klein schreiben)
      - Substantiven (groß schreiben)
      - Titel-Regel: erstes & letztes Wort immer groß
    """
    # 1. Zusammengesetzte Begriffe schützen
    protected, composites = _protect_composites(title)

    # 2. SpaCy-Dokument erzeugen
    doc = nlp_de(protected)

    tokens = []
    for tok in doc:
        text, pos = tok.text, tok.pos_
        wl = text.lower()

        # a) SPECIAL_WORDS → feste Schreibweise
        if wl in SPECIAL_WORDS:
            tokens.append(SPECIAL_WORDS[wl])
            continue

        # b) Kurzwörter (Whitelist) → groß schreiben
        if _SHORT_RE.fullmatch(wl):
            tokens.append(wl.upper())
            continue

        # c) Stopwörter → klein
        if _STOP_RE.fullmatch(wl):
            tokens.append(wl)
            continue

        # d) Substantive → groß
        if pos in {"NOUN", "PROPN"}:
            tokens.append(text.capitalize())
            continue

        # e) Adjektive / Adverbien → klein
        if pos in {"ADJ", "ADV"}:
            tokens.append(wl)
            continue

        # f) Rest → Original belassen
        tokens.append(text)

    # 3. Titel-Regel: erstes & letztes Wort groß, Platzhalter auslassen
    if tokens:
        if not tokens[0].startswith("COMPPLACEHOLDER"):
            tokens[0]  = tokens[0].capitalize()
        if not tokens[-1].startswith("COMPPLACEHOLDER"):
            tokens[-1] = tokens[-1].capitalize()

    preliminary = " ".join(tokens)

    # 4. Geschützte Begriffe wieder einsetzen
    for placeholder, original in composites:
        preliminary = re.sub(re.escape(placeholder), original, preliminary, flags=re.IGNORECASE)

    return preliminary

def _dprint(*args, **kw):
    """Kleiner Debug-Helper, kann später durch logging ersetzt werden."""
    if DEBUG:
        print(*args, **kw)

def is_english(word: str,
               LANGID_MIN: float = 0.95,
               FASTTEXT_MIN: float = 0.90) -> bool:
    """
    Erkennt englische Einzelwörter.
    1. Sofort-Whitelist (Tech-Lexeme)
    2. langid + fastText mit sehr hohen Konfidenz-Schwellen
    """
    # --- Typ-Sicherheit -------------------------------------------------
    if not isinstance(word, str):
        return False

    word = word.lower().strip()

    # --- 1. Sofort-Whitelist -------------------------------------------
    if word in ENGLISH_TERMS:
        return True

    # --- 2. Statistische Modelle ---------------------------------------
    try:
        lang_lid, conf_lid = langid.classify(word)
        prob_lid = math.exp(conf_lid)          # Log-Likelihood → Wahrscheinlichkeit
    except Exception:
        prob_lid = 0.0
        lang_lid = ""

    try:
        pred_ft, prob_ft = fasttext_model.predict(word, k=1)
        lang_ft = pred_ft[0].replace("__label__", "")
        prob_ft = prob_ft[0]
    except Exception:
        prob_ft = 0.0
        lang_ft = ""

    return (
        lang_lid == "en" and prob_lid >= LANGID_MIN and
        lang_ft == "en" and prob_ft >= FASTTEXT_MIN
    )


def title_based_keyword_filter(kw_tokens, title_tokens, threshhold:float = 0.3):
    """
    Check if at least 30% of keyword tokens are in title tokens or close matches.
    """
    matches = 0
    for token in kw_tokens:
        if token in title_tokens:
            matches += 1
        else:
            # near match: allow 1 char difference for words >4 letters
            for t in title_tokens:
                if len(token) > 4 and SequenceMatcher(None, token, t).ratio() > 0.85:
                    matches += 1
                    break
    return matches / len(kw_tokens) >= threshhold


def to_number(val):
    """
    Convert string with commas/dots/spaces to a float, or 0.0 if invalid.
    """
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    val = re.sub(r"[^\d.]", "", str(val))
    try:
        return float(val)
    except ValueError:
        return 0.0
    
def is_noun(word: str) -> bool:
    """
    Returns True if word is a noun in German or English.
    """
    if not word or len(word) < 2:
        return False
    doc_de = nlp_de(word)
    # Check POS tags (NOUN = common noun, PROPN = proper noun)
    return any(t.pos_ in {"NOUN", "PROPN"} for t in doc_de)
           

def keyword_similarity_tokens(a: str, b: str, min_len: int = 3) -> float:
    """
    Word-level Jaccard similarity; ignore words shorter than min_len.
    """
    a_tok = {w for w in a.split() if len(w) >= min_len}
    b_tok = {w for w in b.split() if len(w) >= min_len}
    if not a_tok or not b_tok:
        return 0.0
    return len(a_tok & b_tok) / len(a_tok | b_tok)

def token_match_score(kw_tokens, title_tokens):
    if not kw_tokens:
        return 0.0
    # fraction of keyword tokens that are in the title
    return sum(token in title_tokens for token in kw_tokens) / len(kw_tokens)


# --- Main Matching ---
def find_best_keywords(title: str, content: str, df: pd.DataFrame, threshold: float = 0.4, title_weight: float = 0.8, 
                       title_match_weight: float = 0.6, top_n: int = 5, debug: bool = False, return_keywords_only=True):
    title_norm = normalize_text(title)
    content_norm = normalize_text(content)

    # Map search volume to numeric
    volume_map = {
        "very low": 0.0,
        "low": 0.33,
        "medium": 0.66,
        "high": 1.0
    }
    
    
    # Clean & normalize results column
    df["# Suchergebnisse mein Now"] = df["# Suchergebnisse mein Now"].apply(to_number)
    max_results = df["# Suchergebnisse mein Now"].max()

    candidates = []
    
    title_tokens = set(title_norm.split()) 
    
    for _, row in df.iterrows():
        kw = row['Suchwort']
        kw_norm = normalize_text(kw)
        kw_tokens = kw_norm.split()
        
        if not is_meaningful_keyword(kw):
            if len(kw) > 4 and debug: 
                print(f"aussortiertes Keywords mit mehr als vier Buchstaben: {kw}")
            continue  # Skip filler / weak keywords

        # Skip if keyword isn't mostly made from title words (60%)
        if not title_based_keyword_filter(kw_tokens, title_tokens):
            continue
        
      #  if not is_noun(kw):
      #      if DEBUG:
      #          print(f"❌ skipping non-noun keyword: {kw}")
      #      continue

        sim_title = keyword_similarity_tokens(kw_norm, title_norm)
        sim_content = keyword_similarity_tokens(kw_norm, content_norm)

        # Combine with previous similarity logic
        match_score = title_weight * sim_title + (1 - title_weight) * sim_content
        
        
        # Fallback: token overlap only if similarity score is low, and the keyword for the title is searched
        if match_score < threshold:
            token_score = token_match_score(kw_tokens, title_tokens)
            match_score = max(match_score, token_score)
        

        if kw_norm in title_norm:
            match_score = min(match_score + 0.5, 1.0)

        # Fallback: char similarity
        if match_score < threshold:
            char_sim = SequenceMatcher(None, kw_norm, title_norm).ratio()
            match_score = max(match_score, char_sim)

        if match_score >= threshold:
            vol_score = volume_map.get(str(row.get('Kategorie Suchvolumen', '')).lower(), 0.0)
            results_score = row.get("# Suchergebnisse mein Now", 0) / max_results if max_results else 0.0
            
            total_score = (
                title_match_weight * match_score +
                0.1 * vol_score +
                (0.9 - title_match_weight) * results_score
            )
            
            candidates.append((kw, match_score, vol_score, results_score, total_score))

            if debug:
                print(f"🔍 {kw}")
                print(f"   Match: {match_score:.2f}, Volume: {vol_score:.2f}, Results: {results_score:.2f}, Total: {total_score:.2f}")

    # Sort by total_score, then match_score
    candidates.sort(key=lambda x: (x[4], x[1]), reverse=True)
    
    if return_keywords_only:
        return [kw for kw, _, _, _, _ in candidates[:top_n]]
    else:
        return candidates[:top_n]


def is_meaningful_keyword(kw):
    """Filtert sinnlose oder generische Keywords aus."""
    kw_clean = normalize_text(kw)
    if kw_clean in SHORT_WORD_WHITELIST:
        return True
    if kw_clean in STOPWORDS:
        return False
    if len(kw_clean) < 4:
        return False
    return True

def extract_thematic_terms(titel: str, base_keywords: List[str]) -> List[str]:
    """Extrahiert aus Titel und Basis-Keywords thematische Kernbegriffe."""
    core_terms = set()
    for text in [titel] + base_keywords:
        doc = nlp_de(text.lower())
        for token in doc:
            if token.is_alpha and len(token.lemma_) > 3:
                core_terms.add(token.lemma_)
    return list(core_terms)

def semantic_thematic_boost(keyword: str, thematic_terms: List[str]) -> float:
    """Berechnet semantische Nähe zu thematischen Begriffen (Boost 0.8–1.2)."""
    doc_kw = nlp_de(keyword.lower())
    if not doc_kw.vector_norm:
        return 1.0
    similarities = []
    for term in thematic_terms:
        doc_term = nlp_de(term.lower())
        if not doc_term.vector_norm:
            continue
        sim = doc_kw.similarity(doc_term)
        similarities.append(sim)
    if not similarities:
        return 1.0
    avg_sim = sum(similarities) / len(similarities)
    if avg_sim > 0.7:
        return 1.2
    elif avg_sim > 0.55:
        return 1.1
    elif avg_sim < 0.35:
        return 0.9
    else:
        return 1.0

# === Hauptfunktion ---------------------------------------------------------

def filter_top_keywords_from_list(
    titel: str,
    base_keywords: List[str],
    keywords_a: List[str],
    df_b: pd.DataFrame,
    *,
    top_n: int = 30,
    weight_results: float = 0.5,
    weight_match: float = 0.3,
    match_threshold: float = 0.7,
    redundancy_threshold: float = 0.85
) -> str:
    """
    Filtert und priorisiert Keywords aus Keywordliste A anhand von Keyword-Daten aus B.
    Kombiniert semantische Relevanz, Volumen, Ergebnisanzahl und Thematik.
    """
    if df_b.empty:
        return ""

    # --- Gewichtsanpassung
    weight_vol = 1 - weight_results - weight_match
    if weight_vol < 0:
        raise ValueError("Summe aus weight_results + weight_match darf nicht > 1 sein.")

    # --- Daten vorbereiten
    df = df_b.copy()
    df["norm"] = df["Suchwort"].apply(normalize_text)
    df["# Suchergebnisse mein Now"] = df["# Suchergebnisse mein Now"].apply(
        lambda x: float(re.sub(r"[^\d.]", "", str(x))) if pd.notna(x) else 0.0
    )

    # --- Scores aus vorhandenen Daten
    max_res = df["# Suchergebnisse mein Now"].max() or 1.0
    df["res_score"] = df["# Suchergebnisse mein Now"] / max_res

    vol_map = {"very low": 0.0, "low": 0.33, "medium": 0.66, "high": 1.0}
    df["vol_score"] = (
        df["Kategorie Suchvolumen"].astype(str).str.lower().map(vol_map).fillna(0.0)
    )

    # --- Thematische Begriffe extrahieren
    thematic_bias = extract_thematic_terms(titel, base_keywords)

    # --- Semantischen Boost berechnen
    df["thema_boost"] = df["Suchwort"].apply(
        lambda kw: semantic_thematic_boost(kw, thematic_bias)
    )

    # --- Gesamtbewertung kombinieren
    df["score"] = (
        weight_vol * df["vol_score"]
        + weight_results * df["res_score"]
    ) * df["thema_boost"]

    # --- Matching Keywords aus A zu B suchen
    picked = []
    for kw_a in keywords_a:
        norm_a = normalize_text(kw_a)
        best_idx = None
        best_sim = 0.0
        for idx, norm_b in df["norm"].items():
            sim = keyword_similarity_tokens(norm_a, norm_b)
            if sim > best_sim:
                best_sim = sim
                best_idx = idx
        if best_idx is not None and best_sim >= match_threshold:
            picked.append({
                "idx": best_idx,
                "keyword": df.at[best_idx, "Suchwort"],
                "score": df.at[best_idx, "score"],
                "sim": best_sim,
            })

    # --- Priorität berechnen
    for p in picked:
        vol_score = df.at[p["idx"], "vol_score"]
        p["priority"] = (
            weight_match * p["sim"]
            + weight_results * p["score"]
            + weight_vol * vol_score
        )

    # --- Sortierung nach Priorität
    candidates = (
        pd.DataFrame(picked)
        .sort_values("priority", ascending=False)
        .drop_duplicates(subset=["keyword"])
    )

    # --- Redundanzfilter
    seen_norm = []
    filtered_keywords = []
    for _, row in candidates.iterrows():
        kw = row["keyword"]
        norm_kw = normalize_text(kw)
        if any(keyword_similarity_tokens(norm_kw, s) > redundancy_threshold for s in seen_norm):
            continue
        if not is_meaningful_keyword(kw):
            continue
        seen_norm.append(norm_kw)
        filtered_keywords.append(kw)
        if len(filtered_keywords) >= top_n:
            break

    # --- Auffüllen mit A, falls weniger als top_n
    if len(filtered_keywords) < top_n:
        used_norms = {normalize_text(kw) for kw in filtered_keywords}
        for kw_a in keywords_a:
            norm_a = normalize_text(kw_a)
            if norm_a in used_norms or not is_meaningful_keyword(kw_a):
                continue
            if any(keyword_similarity_tokens(norm_a, s) > redundancy_threshold for s in used_norms):
                continue
            filtered_keywords.append(kw_a)
            used_norms.add(norm_a)
            if len(filtered_keywords) >= top_n:
                break

    return ", ".join(filtered_keywords[:top_n])


def dedup_candidates(candidates: List[Tuple[str, float, float, float, float]],
                     sim_cut: float = 0.8) -> List[Tuple[str, float, float, float, float]]:
    """
    Entfernt „ähnliche“ Keywords aus einer Liste von 5-Tupeln.
    Reihenfolge bleibt erhalten (bestes steht vorne).
    Rückgabe: Liste von 5-Tupeln, aber ohne Doppelungen.
    """
    if not candidates:
        return []

    kept = []                      # finale Liste
    seen = []                      # nur die normalisierten „Stemm“-Strings

    for kw, m, v, r, t in candidates:
        stem = kw.lower().replace("-", " ").replace("+", " plus").strip()
        duplicate = False

        for prev_stem in seen:
            if SequenceMatcher(None, stem, prev_stem).ratio() >= sim_cut:
                duplicate = True
                break

        if not duplicate:
            kept.append((kw, m, v, r, t))
            seen.append(stem)

    return kept


# ========== KI Prompt

def run_ai_prompt(
    key: str,
    pm: PromptManager,
    model: str = "gpt-4o",
    max_output_tokens: int = 300,
    **placeholders
) -> str:
    """
    Führt einen gespeicherten Prompt mit variablen Platzhaltern aus.
    Nutzt die gespeicherte Temperatur & JSON-Template aus PromptManager.
    """

    data = pm[key]  # enthält {prompt: "...", temperature: ..}

    if not isinstance(data, dict) or "prompt" not in data:
        raise ValueError(f"Prompt '{key}' fehlt oder ist nicht korrekt gespeichert.")

    template = data["prompt"]
    temperature = data.get("temperature", 0.5)

    # Platzhalter einsetzen:
    try:
        prompt = template.format(**placeholders)
    except KeyError as missing:
        raise KeyError(f"Fehlender Prompt-Platzhalter: {missing} in '{key}'")

    resp = client.responses.create(
        model=model,
        input=prompt,
        temperature=temperature,
        max_output_tokens=max_output_tokens
    )

    return resp.output_text.strip()


def ai_abschlussart(clean_abschluss_field: str, worker) -> str:
    """Kombiniert semantisch vollständige und zeichenbegrenzte Bereinigung."""
    pm5 = PromptManager(initial_data=prompts)
    result = SAFE(run_ai_prompt, "ai_abschlussart_primary", pm5, clean_abschluss_field=clean_abschluss_field, model="gpt-4o", max_output_tokens=200, logger=worker.error.emit)
    if len(result) > 250:
        pm6 = PromptManager(initial_data=prompts)
        print(f"Hinweis: Text >250 Zeichen ({len(result)}). Kürze über zweiten Prompt.")
        result = SAFE(run_ai_prompt, "enforce_length_via_ai", pm6, long_text=result, length=len(result), model="gpt-4o", max_output_tokens=200, logger=worker.error.emit)
    return result


def clean_abschluss_fields(text, singular, plural):
    """
    Vereinheitlicht Abschlussbezeichnungen:
    - Mehrfach vorkommender Singular → einmal Plural
    - Singular + Plural gemischt → einmal Plural
    - Einmaliger Singular → bleibt Singular
    """
    if not isinstance(text, str):
        return text

    # Alle Varianten vereinheitlichen (Bindestrich, Plural)
    text_norm = re.sub(rf"\b{re.escape(plural)}\b", singular, text)
    
    # Zähle jetzt die vereinheitlichte Singularform
    count = text_norm.count(singular)

    if count > 1:
        # Alle Vorkommen des Singular entfernen
        text_norm = re.sub(rf"(,?\s*){re.escape(singular)}", "", text_norm)
        # Einmal den Plural anhängen
        if text_norm.strip():
            text_norm += f", {plural}"
        else:
            text_norm = plural

    # Aufräumen: doppelte Kommata und Leerzeichen
    text_norm = re.sub(r',\s*,', ',', text_norm).strip(', ')

    return text_norm

def process_abschlussfeld(
    df, idx, feld_liste, feldname, singular_label, plural_label,
    ki_function, worker, min_length_for_ki=200, max_kurse=None
):
    """
    Vereinheitlicht Abschlussfelder, ruft bei Bedarf die KI auf und schreibt Ergebnis ins DataFrame.
    KI wird nur aufgerufen, wenn die bereinigte Länge >= min_length_for_ki.
    """
    # --- Zusammenführen & Bereinigen ---
    joined_field = ", ".join(filter(None, map(str, feld_liste))).strip(", ")

    clean_field = clean_abschluss_fields(
        joined_field,
        singular_label,
        plural_label
    )

    # --- Prüfen, ob KI-Aufruf nötig ---
    if (
        clean_field
        and clean_field.strip()
        and (max_kurse is None or idx < max_kurse)
        and len(clean_field) >= min_length_for_ki
    ):
        try:
            ai_output = _as_str(SAFE(ki_function, clean_field, worker=worker))
            final_output = (
                f"[KI-Zusammenfassung]\n{ai_output}\n\n[Original-Inhalte]\n{clean_field}"
            )
        except Exception as e:
            print(f"Warnung: KI-Bereinigung für {feldname} fehlgeschlagen ({e})")
            final_output = clean_field
    else:
        final_output = clean_field

    df.at[idx, feldname] = final_output


def safe_save(wb, filename, timeout=20):
    """
    Thread-safe Excel-Save mit Timeout:
    - Speichert zuerst temporär (filename + ".tmp")
    - Überschreibt Original nur bei erfolgreichem Save
    - Timeout beendet hängende Saves
    """
    tmp_filename = filename + ".tmp"
    # print(f"💾 safe_save start: {filename}")

    result = {}

    def _do_save():
        try:
            wb.save(tmp_filename)  # kann hängen
            result["ok"] = True
        except Exception as e:
            result["error"] = e

    # --- SAVE OHNE LOCK (kann parallel laufen, braucht kein Excel-Zugriff) ---
    t = threading.Thread(target=_do_save, daemon=True)
    t.start()
    t.join(timeout)

    if t.is_alive():
        print(f"⚠️ Save timeout nach {timeout}s – breche Speichern ab")
        # Thread läuft weiter, aber wir verlassen die Funktion
        if os.path.exists(tmp_filename):
            try:
                os.remove(tmp_filename)
                print(f"🗑️ Temporäre Datei gelöscht (Timeout): {tmp_filename}")
            except Exception as e:
                print(f"⚠️ Konnte temporäre Datei nicht löschen: {e}")
        return False

    # --- LOCK NUR BEIM MOVE (kritischer Bereich) ---
    with write_lock:
        if "error" in result:
            print(f"⚠️ Speichern fehlgeschlagen: {result['error']}")
            if os.path.exists(tmp_filename):
                try:
                    os.remove(tmp_filename)
                    print(f"🗑️ Temporäre Datei gelöscht (Fehler): {tmp_filename}")
                except Exception as e:
                    print(f"⚠️ Konnte temporäre Datei nicht löschen: {e}")
            return False

        if os.path.exists(tmp_filename):
            try:
                shutil.move(tmp_filename, filename)
                print(f"✅ Erfolgreich gespeichert: {filename}")
                return True
            except Exception as e:
                print(f"⚠️ Fehler beim Move: {e}")
                return False
        else:
            print("⚠️ Keine temporäre Datei gefunden, nichts gespeichert.")
            return False


def safe_call(func, *args, model="gpt-4o", retries=3, wait=2, consume_tokens=None, logger=None,**kwargs):
    """
    - Token-Budget prüfen (thread-safe via 'consume_tokens')
    - Exponentielles Backoff bei Fehlern
    """
    # Eingabe für Token-Schätzung
    parts = [str(x) for x in args if x is not None]
    kw_parts = [f"{k}:{v}" for k, v in kwargs.items() if v is not None]
    text_input = " ".join(parts + kw_parts)

    if consume_tokens is not None:
        consume_tokens(text_input, model=model)

    for attempt in range(1, retries + 1):
        try:
            return func(*args, **kwargs)
        except KeyboardInterrupt:
            raise
        except Exception as e:
            print(f"⚠️ Error in safe_call: {e} (retry {attempt}/{retries})")
            if logger:
                logger(f"safe_call Fehler: {e}\n\n{traceback.format_exc()}")
            if attempt < retries:
                time.sleep(wait * attempt)  # exponentielles Backoff
            else:
                raise RuntimeError("❌ Max retries exceeded in safe_call")

    
# --- Token-Schätzer (robust, mit Fallback) ---
def _estimate_tokens(text: str, model="gpt-4o") -> int:
    try:
        import tiktoken
        try:
            enc = tiktoken.encoding_for_model(model)
        except Exception:
            # robuster Fallback für GPT-4/4o
            enc = tiktoken.get_encoding("cl100k_base")
        return len(enc.encode(text or ""))
    except Exception:
        # sehr grobe Heuristik als letzter Fallback
        # (≈4 Zeichen pro Token)
        return max(1, int(len((text or "")) / 4))


def make_token_consumer(max_tokens_per_min: int):
    state = {
        "used": 0,
        "reset": time.time() + 90.0,
    }
    lock = threading.Lock()

    def consume(text: str, model="gpt-4o") -> int:
        tokens_needed = _estimate_tokens(text, model)
        while True:
            with lock:
                now = time.time()
                if now > state["reset"]:
                    state["used"] = 0
                    state["reset"] = now + 90.0

                if state["used"] + tokens_needed <= max_tokens_per_min:
                    state["used"] += tokens_needed
                    return tokens_needed

                wait_for = max(0.0, state["reset"] - now)
            # außerhalb des Locks schlafen
            print(f"⏳ Token budget exceeded. Waiting {wait_for:.1f} sec...")
            time.sleep(wait_for + 0.1)
    return consume


# ======= definiere token_consume
token_consume = make_token_consumer(MAX_TOKENS_PER_MIN)

        
# ========== definiere Standard Funktionsaufruf    
SAFE = partial(safe_call, consume_tokens=token_consume, model="gpt-4o")
    

def _norm(text: str) -> List[str]:
    """Normalisiert und filtert sinnvolle Keywords"""
    if pd.isna(text):
        return []
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ASCII", "ignore").decode()
    words = re.findall(r"\w+", text)
    # Filterung mit is_meaningful_keyword
    return [w.lower() for w in words if is_meaningful_keyword(w)]

# 2. Flache Systematik-Liste erzeugen (einmalig)
SYSTEMATIK_FLAT = None

def expand_semantic_terms_spacy(
    terms: List[str],
    thresholds: list[float] = [0.75, 0.65]
) -> Set[str]:
    """
    Erweitert Begriffe semantisch aus SYSTEMATIK_FLAT ohne Dopplungen.
    Nutzt mehrere Thresholds für mehr/flexiblere Treffer.
    """
    expanded = set(terms)
    
    # Nur lange, sinnvolle Begriffe prozessieren
    terms_docs = [(t, nlp_de(t)) for t in set(terms) if len(t) > 3 and t in SYSTEMATIK_FLAT]
    sys_docs = [(s, nlp_de(s)) for s in SYSTEMATIK_FLAT if len(s) > 3]
    
    for term_text, term_doc in terms_docs:
        best_matches = []
        for sys_text, sys_doc in sys_docs:
            # Skip if vectors missing → avoid W008
            if (term_doc.has_vector is False) or (sys_doc.has_vector is False):
                continue
            sim = term_doc.similarity(sys_doc)
            if sim > thresholds[0]:  # Hoher Threshold = direkt hinzufügen
                expanded.add(sys_text)
            elif sim > thresholds[1]:  # Mittlerer Threshold = als Kandidat
                best_matches.append((sim, sys_text))
        
        # Top 2 Kandidaten aus mittlerem Bereich auch hinzunehmen
        best_matches.sort(reverse=True)
        for _, sys_text in best_matches[:2]:
            expanded.add(sys_text)
    
    return expanded

def systematik_vorauswahl(
    titel: str,
    basis: str,
    inhalt: str,
    top: int = 12,
    gewichtung_titel: int = 5,
    max_zusatz_woerter: int = 8,
    semantic_weight: float = 0.8
) -> List[str]:
    """Hauptfunktion mit semantischer Expansion und dedupliziertem Scoring"""

    # === 1. Normalisierung + Deduplizierung ===
    titel_woerter = _norm(titel)
    basis_woerter = _norm(basis)
    inhalt_woerter = _norm(inhalt)
    
    # === 2. SEMANTISCHE EXPANSION (nur der wichtigsten Begriffe) ===
    # Expansion der Top 8 Begriffe (Titel + Basis) für Performance
    top_terms = list(dict.fromkeys(titel_woerter[:4] + basis_woerter[:4]))
    semantic_set = expand_semantic_terms_spacy(top_terms)
    
    # === 3. GEWICHTETER COUNTER (statt Listen-Wiederholung) ===
    pool_counter = Counter()
    
    # Titelgewichtung (exakt, ohne Dopplungen)
    for wort in titel_woerter:
        pool_counter[wort] += gewichtung_titel
    
    # Basis
    for wort in basis_woerter:
        pool_counter[wort] += 1.0
    
    # Semantische Erweiterungen mit reduzierter Gewichtung
    for wort in semantic_set:
        if wort not in pool_counter:  # Nur wenn neu
            pool_counter[wort] += semantic_weight
    
    # Inhalt: Häufigste neue Wörter, die nicht im Titel sind
    titel_set = set(titel_woerter)
    inhalt_zusatz = [w for w in inhalt_woerter if w not in titel_set]
    for wort, _ in Counter(inhalt_zusatz).most_common(max_zusatz_woerter):
        pool_counter[wort] += 0.5
    
    # Debug-Ausgabe
    #print(f"Pool-Wörter ({len(pool_counter)}): {dict(list(pool_counter.items())[:10])}")
    
    # === 4. SCORING gegen SYSTEMATIK ===
    scored = []
    for systematik in SYSTEMATIK:
        sys_norm = set(_norm(systematik))
        
        # Exakte Übereinstimmungen
        base_score = sum(pool_counter[w] for w in sys_norm if w in pool_counter)
        
        # Semantischer Bonus: wenn ein Systematik-Wort in unserer semantischen Erweiterung ist
        semantic_bonus = sum(
            semantic_weight for w in sys_norm 
            if w in semantic_set and w not in pool_counter
        )
        
        total_score = base_score + semantic_bonus
        if total_score > 0:  # Nur relevante Systematiken
            scored.append((total_score, systematik))
    
    # === 5. Filter & Return ===
    scored.sort(key=lambda x: x[0], reverse=True)
    min_score = max(1, gewichtung_titel // 3)  # Dynamische Schwelle
    
    result = [s for score, s in scored[:top] if score >= min_score]
    return result if result else ["keine Systematik gefunden"]



# ---------- LLM ----------
def extract_inhalte_auf_einen_blick(text):
    """
    Extrahiert den Abschnitt 'Inhalte auf einen Blick' 
    - nur die erste Zeile mit den kommagetrennten Inhalten
    """
    pattern = r"^\s*!Inhalte auf einen Blick:\s*(.+?)$"
    match = re.search(pattern, text, re.MULTILINE | re.IGNORECASE)
    
    return match.group(1).strip() if match else None
    
    return None    
    
WORD_SPLIT = re.compile(r'[\s/\\|,:;–—-]+')   # auch Bindestrich trennt!

def split_title(title: str) -> list[str]:
    """Zerteilt Titel und liefert Wörter ≥ 4 ODER in SHORT_WORD_WHITELIST."""
    return [
        w.strip()
        for w in WORD_SPLIT.split(title)
        if len(w.strip()) >= 4 or w.strip().lower() in SHORT_WORD_WHITELIST
    ]



def collect_candidates_debug(titles: list[str], df: pd.DataFrame) -> list[str]:
    """Gibt ALLE Keywords zurück, deren normiertes Wort im Titel vorkommt – ohne Filter!"""
    found = []
    for ttl in titles:
        if not isinstance(ttl, str):
            print(f"❌ ttl ist kein str: {ttl!r} (Typ: {type(ttl)})")
            continue
        ttl_norm = normalize_text(ttl)
        for _, row in df.iterrows():
            kw_norm = normalize_text(row['Suchwort'])
            if kw_norm in ttl_norm or ttl_norm in kw_norm:
                found.append(row['Suchwort'])
    return list(dict.fromkeys(found))   # Duplikate entfernen, Reihenfolge behalten

def collect_candidates(titles: List[str],
                       content: str,
                       df: pd.DataFrame,
                       threshold: float = 0.4,
                       title_weight: float = 0.8,
                       title_match_weight: float = 0.6) -> List[Tuple]:
    """Ruft für jede Teilüberschrift find_best_keywords(..., return_keywords_only=False) auf
       und sammelt ALLE Ergebnisse in EINER Liste."""
    all_cand = []
    for t in titles:
        all_cand.extend(
            find_best_keywords(t, content, df, threshold, title_weight,
                               title_match_weight, top_n=30, return_keywords_only=False)
        )
    return all_cand


def similar_enough(a, b, threshold=0.92):
    return SequenceMatcher(None, normalize_text(_as_str(a)), normalize_text(_as_str(b))).ratio() >= threshold


def apply_basic_subs(title: str) -> str:
    # --- 1. Fix: PRINCE2 ---
    title = re.sub(r"(?i)prince\s*2(\s*®)?", "PRINCE2 ®", title)
    
    # --- 2. Entfernen / Ersetzen von Berufsbezeichnungen ---
    replacement = random.choice(["Experte", "Spezialist"])
    title = re.sub(r"\bFach(kräfte|kraft|leute|mann|frau)\b", replacement, title, flags=re.IGNORECASE)
    
    # --- 3. Gender-Endungen löschen ---
    title = re.sub(r"(\*in(nen)?|:in(nen)?|/-?in(nen)?|-in(nen)?|-frau(en)?)\b", "", title, flags=re.IGNORECASE)
    
    # --- 4. Spezifische Ersetzungen ---
    title = re.sub(r"Assisten(t|tin|tinnen|ten|t:in|t\-in|t\*in)?", "Assistenz", title, flags=re.IGNORECASE)
    title = re.sub(r"Projektleiter(innen|in|:in|/-in|\*in)?", "Projektleiter", title, flags=re.IGNORECASE)
    title = re.sub(r"Zertifizierung(sprogramm|skurs|slehrgang|smodul)?", "Zertifizierung", title, flags=re.IGNORECASE)
    
    # --- 4.1: Mehrfache „z“ in Assistenz entfernen ---
    title = re.sub(r"Assistenz{2,}", "Assistenz", title)
    

    
    # --- 5.1. Alle Anführungszeichen entfernen ---
    title = re.sub(r"['\"“”„‟‚‘‛❛❜❝❞❮❯‹›«»`´]", "", title)
    
    # --- 6. Satzzeichen-Spationierung ---
    title = re.sub(r"\s+([,.;])", r"\1", title)
    title = re.sub(r"([,.;])([^\s])", r"\1 \2", title)
    
    # --- 7. Bindestriche vereinheitlichen ---
    title = re.sub(r"\s*-\s*", "-", title)
    title = re.sub(r"\b-und\b", "- und", title)
    title = re.sub(r"\s*–\s*", " – ", title)
    
    # --- 8. Slash- und SAP/HANA-Regeln (sicher & erweitert) ---
    # Leerzeichen rund um Slashes bereinigen
    title = re.sub(r"\s*/\s*", "/", title)
    # Slashes am Ende entfernen
    title = re.sub(r"/\s*$", "", title.strip())
    
    # S/4HANA vereinheitlichen (verschiedene Schreibweisen)
    title = re.sub(r"(?i)s\s*/?\s*4\s*hana", "S/4HANA", title, flags=re.IGNORECASE)
    
    # Wenn SAP und HANA beide im Text vorkommen → nur diese Begriffe ersetzen, Rest bleibt
    if re.search(r"(?i)\bSAP\b", title) and re.search(r"(?i)\bHANA\b", title):
        title = re.sub(r"(?i)\bSAP\b", "SAP S/4HANA", title)
        title = re.sub(r"(?i)\bHANA\b", "", title)
        # Doppelte Leerzeichen entfernen nach der Ersetzung
        title = re.sub(r"\s+", " ", title).strip()
        
    # --- SAP-Schreibweise vereinheitlichen ---
    title = re.sub(r"(?i)sap", "SAP", title)
    
    # --- 9. Klammern kompakt ---
    title = re.sub(r"\(\s+", "(", title)
    title = re.sub(r"\s+\)", ")", title)
    
    # --- 10. Mehrfache Leerzeichen bereinigen ---
    title = re.sub(r"\s+", " ", title).strip()
    
    # --- 4.2: Nach jedem beliebigen Wort-Buchstaben + Bindestrich
    #          den folgenden Buchstaben großschreiben
    title = re.sub(
        r"(\w)-([a-zäöüß])",                 # 1. Buchstabe-Ziffer_unterstrich + - + Kleinbuchstabe
        lambda m: m.group(1) + "-" + m.group(2).upper(),
        title,
        flags=re.IGNORECASE
    )
    
    return title

    
def create_title(split_parts: list[str], best_kw: str, worker) -> List[str]:
    """
    Erzeugt optimierte Varianten eines Weiterbildungstitels.
    Eingabe: split_parts (erstes Element = Original-Titel, Rest = aufgesplittete Teile)
    Ausgabe: Einzelner neuer Titel (str) – entweder übersetzt (wenn 50 %+ englisch) oder Varianten mit best_kw.
    """

    # --- 1. Feedback-Loop: Englisch-Erkennung über split_parts -----------------
    original_title = split_parts[0]          # Original bleibt unverändert
    candidate_words = split_parts[1:]        # Restliche Teile

    if not candidate_words:                  # keine Teile -> sofort Deutsch-Varianten
        english_ratio = 0.0
    else:
        eng_count = sum(1 for w in candidate_words if is_english(w))
        english_ratio = eng_count / len(candidate_words)
    if DEBUG:
        print(f"Englischanteil: {english_ratio}")

    # 50 %-Schwelle
    if english_ratio >= 0.5:
        # --- 2. Als englisch erkannt -> komplett übersetzen -------------------
        if DEBUG:
            print(f"Titel '{original_title}' als Englisch erkannt (mehr als 50% Englischanteil)")
            log(f"Titel '{original_title}' als Englisch erkannt (mehr als 50% Englischanteil)")
        pm10 = PromptManager(initial_data=prompts)
        translated = SAFE(run_ai_prompt, "translate_title_to_german", pm10, title=original_title, model="gpt-4o", max_output_tokens=150, logger=worker.error.emit)
        # --- KI-Ausgabe normieren --------------------------
        raw_lines = translated.strip().splitlines()
        ki_variants = [line.strip() for line in raw_lines if line.strip()]
        ki_variants = list(dict.fromkeys(ki_variants))   # Duplikate entfernen
        return ki_variants
    if DEBUG:
        print(f"Titel '{original_title}' nicht als Englisch erkannt. ")

    # --- 3. Sonst: Varianten mit best_kw erzeugen (wie bisher) ---------------
    title_clean   = apply_basic_subs(original_title)
    best_kw_clean = apply_basic_subs(best_kw)

    variants = [
        f"{best_kw_clean} – {title_clean}",
        f"{title_clean} – {best_kw_clean}",
        title_clean
    ]

    final_variants = []
    for v in variants:
        v_norm = normalize_with_spacy(v)
        v_norm = apply_basic_subs(v_norm)
        final_variants.append(v_norm.strip())

    # Duplikate entfernen, Reihenfolge behalten
    final_variants = list(dict.fromkeys(final_variants))

    return final_variants

def extract_titles_from_inhalt(inhalt: str) -> str:
    """
    - Erste Zeile des Feldes (sofern mit '!' beginnend) gilt als Titel.
    - Danach jede Zeile, die direkt auf eine LEERE Zeile folgt und mit '!' beginnt.
    Rückgabe: Zeilenweise nummerierte Titel (ohne '!'), leere Zeilen werden ignoriert.
    """
    titles = []
    lines = inhalt.splitlines()

    # 1) Erste Zeile prüfen
    if lines and lines[0].strip().startswith("!"):
        titles.append(lines[0].strip()[1:].strip())

    # 2) Nach Leerzeile + '!'
    for prev, curr in zip(lines, lines[1:]):
        if prev.strip() == "" and curr.strip().startswith("!"):
            titles.append(curr.strip()[1:].strip())

    # 3) Nummerierte Ausgabe
    return "\n".join(f"{i}. {t}" for i, t in enumerate(titles, start=1)) if titles else ""



def _as_str1(x: object) -> str:
    # convert anything (incl. tuples/lists/NA) into clean string
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    if isinstance(x, (list, tuple)):
        x = " ".join(_as_str(v) for v in x)
    return str(x)

def _as_str(x):
    if isinstance(x, (list, tuple, np.ndarray, pd.Series)):
        return " ".join(map(str, x))
    return _as_str1(x)

def is_scalar_na(x):
    return (x is None) or (isinstance(x, float) and pd.isna(x))

def sanitised_mod_ids(raw: str) -> list[str]:
    ids = re.findall(r"[A-Z]?\d+[-/]?\d+", str(raw))
    return [m.strip().upper() for m in ids if m.strip()]





unfound_modules = []

start = time.time()

def process_row(idx, row, modulkombi_df, ws, wb, backup_file, stop_event, options, worker):
    
    if stop_event.is_set():
        # print(f"⏹️ Thread {idx + 1} abgebrochen vor Verarbeitung")
        return
    
    # if DEBUG:
    #     print("current_row_title raw:", row.get(Kursübersicht, "Unbekannt"))

    _row_copy = row.copy()
    retries = 0
    while retries <= MAX_RETRIES:
        try:
            if MaxRows is not None and idx > MaxRows:
                if DEBUG: 
                    print(f"Stopp nach Reihe: {idx-1}")
                return
    
            mod_str = row[Modulnummern]
            if pd.isna(mod_str):
                return
    
            # -------- helpers (robust to NA/tuples) --------    
            def _clean_html(x: object) -> str:
                return html_to_clean_text(_as_str(x)).strip()
    
            current_row_title = _as_str(row[Kursübersicht]) if Kursübersicht in row else "Unbekannt"
    
    
            mod_ids = [m.strip().upper() for m in _as_str(mod_str).split(",") if m.strip()]
            
            if idx > 0:
                log(f"✅ Zeile {idx} abgeschlossen")
            print(f"🔎  Zeile {idx + 1}: {mod_str}")
            log(f"🔎  Zeile {idx + 1}: {mod_str}")
        
            matched_modules = find_module_rows(worker, module_ids=mod_ids, kurs_title=current_row_title, idx=idx,  debug=False)
            
            print("Gefundene Module:")
            for n in matched_modules["Modulnummer2"]:
                print(" -", n)
            
            log("Gefundene Module:\n" + "\n".join(f" - {n}" for n in matched_modules["Modulnummer2"]))

            


            
            if not matched_modules.empty:
                inhalte, zielgruppen, voraussetzungen = [], [], []
                abschlussart, abschlussbezeichnung = [], []
    
                for mr in matched_modules.itertuples(index=False):
                    titel = _as_str(getattr(mr, "Titel", "")).strip()
    
                    # Inhalte
                    inhalt_raw = getattr(mr, "Bildungsinhalt", "")
                    inhalt_clean = strip_intro(_clean_html(inhalt_raw))
                    if inhalt_clean is not None:
                        inhalte.append(f"!{titel}\n{inhalt_clean}")
    
                    # Zielgruppe
                    ziel = _clean_html(getattr(mr, "Zielgruppe", ""))
                    if ziel is not None:
                        zielgruppen.append(ziel)
    
                    # Voraussetzungen
                    voraus = _clean_html(getattr(mr, "Zugang", ""))
                    if voraus is not None:
                        voraussetzungen.append(voraus)
    
                    # Abschlussart/-bezeichnung
                    abs_art = _clean_html(getattr(mr, "Abschlussart", ""))
                    if abs_art is not None:
                        abschlussart.append(abs_art)
    
                    abs_bez = _clean_html(getattr(mr, "Abschlussbezeichnung", ""))
                    if abs_bez is not None:
                        abschlussbezeichnung.append(abs_bez)
                    
    
                original_title = _as_str(row[TitelFeld])
                inhalt_zusammen = "\n\n".join(_as_str(x) for x in inhalte).strip()
                

                
                split_words = split_title(original_title)
                
                split_parts = [original_title] + [w for w in split_words if w not in {original_title}]
                
                AnzahlderZeilen = 3
                
                # Übersetzen aller englischen Wörter und an split_parts anfügen
                for w in split_words:
                    if is_english(w):
                        pm0 = PromptManager(initial_data=prompts)  
                        for line in SAFE(run_ai_prompt, "translate_to_german", pm0, title=w, model="gpt-4o", max_output_tokens=100, logger=worker.error.emit).splitlines()[:AnzahlderZeilen]:
                            t = line.strip()
                            if t and t.lower() not in {p.lower() for p in split_parts}:
                                split_parts.append(t)
                
                # print("Titel-Split-Liste : ", split_parts)
                            
                #  Normale Keyword-Suche (mit Filter)
                raw_candidates = collect_candidates(split_parts, "", keywords_df, threshold=0.2)
                    
                unique_candidates = dedup_candidates(raw_candidates)
                
                    
                debug_df = (pd.DataFrame(unique_candidates,
                                         columns=['keyword', 'match_score', 'volume_score',
                                                  'results_score', 'total_score'])
                            .sort_values(['total_score', 'match_score'], ascending=False)
                            .drop_duplicates(subset=['keyword'], keep='first')
                            .reset_index(drop=True))
                
                best_kw_rows = debug_df[['keyword', 'match_score', 'volume_score',
                                   'results_score', 'total_score']].to_records(index=False).tolist()
    
                # 🔧 Falls keine Keywords gefunden wurden, Dummy-Zeile mit Titel einfügen
                if not best_kw_rows or len(best_kw_rows) == 0:
                    best_kw_rows = [(original_title, 0.0, 0.0, 0.0, 0.0)]            
                    
                if DEBUG and False:
                    print("Titel: ", original_title)
                    for kw, match, vol, results, total in best_kw_rows:
                        print(f"{kw} | Match: {match:.2f} | Volume: {vol:.2f} | Results: {results:.2f} | Total: {total:.2f}")
    
                kursuebersicht_neu = extract_titles_from_inhalt(inhalt_zusammen)
    
                detail_data1 = {
                    'kursuebersicht': kursuebersicht_neu,
                    'original_title': original_title,
                    'inhalte': inhalt_zusammen,
                }
                
                # =========== Best Keyword Feedback Loop =============
                if options.createMasterKeyword and best_kw_rows is not None and len(best_kw_rows) > 0:
                    worker.ask_keyword.emit({
                        "title":original_title,
                        "rows":best_kw_rows,
                        "keywords_df":keywords_df,              # Suche aktivieren
                        "search_threshold":0.3,                 # Mindest-Ähnlichkeit
                        "search_top_n":25,                      # Max Ergebnisse
                        "detail_data":detail_data1
                    })
                    user_kw = worker.wait_for_gui_result()
                    if user_kw is None:
                        # print("🚨 Benutzer hat Abbrechen gedrückt – Fortschritt wird gespeichert ...")
                        stop_event.set()
                        raise UserCancelledError("User pressed Cancel in KeywordSelectionDialog")
                    else:
                        best_kw = user_kw
                else:
                    best_kw = ""
                    
                if DEBUG:
                    print("Haupt-Keyword: ", best_kw)
                    log("Haupt-Keyword: ", best_kw)
                
                if options.createTitel:
                    title_splits = [original_title] + [w for w in split_words if w not in {original_title}]
                    new_titles = create_title(title_splits, best_kw, worker)
                    # print(f"Neue Titel-Varianten Reihe {idx + 1}: ", new_titles)
                    
                    # =========== Titel Feedback Loop =============
                    kursuebersicht_neu = extract_titles_from_inhalt(inhalt_zusammen)
        
                    detail_data2 = {
                        'kursuebersicht': kursuebersicht_neu,
                        'original_title': original_title,
                        'inhalte': inhalt_zusammen,
                        "best_kw": best_kw,
                    }
                    
                    worker.ask_title.emit({
                        "original_title":original_title, 
                        "new_titles":new_titles, 
                        "detail_data":detail_data2})
                    
                    selected_title = worker.wait_for_gui_result()
                    
                    if selected_title is None:
                        # print("🚨 Benutzer hat Abbrechen gedrückt – Fortschritt wird gespeichert ...")
                        stop_event.set()
                        raise UserCancelledError("User pressed Cancel in KeywordSelectionDialog")
                        
        
                    if selected_title:
                        print("Gewählter/bearbeiteter Titel:", selected_title)
                        log("Gewählter/bearbeiteter Titel:", selected_title)
        
                    modulkombi_df.at[idx, TitelNeu] = selected_title if selected_title else original_title
                
    
                if stop_event.is_set():
                    # print(f"⏹️ Thread {idx + 1} abgebrochen vor Verarbeitung")
                    return
                if options.createEinleitung:
                    if DEBUG:
                        print(f"vor Einleitung === Zeile {idx + 1}, jetzt bei {time.time()-start:.2f} sec")
        
                    # --- Einleitung ---
                    if idx < AnzahlKurse:
                        pm = PromptManager(initial_data=prompts)
                        einleitung = SAFE(run_ai_prompt, "ai_einleitung", 
                                          pm=pm, 
                                          titel=original_title, 
                                          inhalte=current_row_title, 
                                          keyword=best_kw, 
                                          max_output_tokens=220, 
                                          model="gpt-4o",
                                          logger=worker.error.emit)
                        modulkombi_df.at[idx, EinleitungNeu] = _as_str(einleitung)
                        if DEBUG:
                            print(f"Einleitung erstellt  === Zeile {idx + 1}, jetzt bei {time.time()-start:.2f} sec")
                            log(f"Einleitung erstellt  === Zeile {idx + 1}")
                    else:
                        einleitung = ""
                else:
                    einleitung = ""
                      
                if stop_event.is_set():
                    # print(f"⏹️ Thread {idx + 1} abgebrochen vor Verarbeitung")
                    return
                # --- Inhalte ---
                if options.createInhalte:
                    pm2 = PromptManager(initial_data=prompts)  
                    if inhalt_zusammen and inhalt_zusammen.strip() != "" and idx < AnzahlKurse:
                        inhalt_ai = _as_str(
                            SAFE(run_ai_prompt, "ai_inhalte",
                                 pm = pm2,
                                 titel=original_title, 
                                 inhalte=current_row_title, 
                                 keyword=best_kw, 
                                 inhaltsfeld=inhalt_zusammen, 
                                 einleitung=einleitung,
                                 max_output_tokens=10000,
                                 model="gpt-4o", logger=worker.error.emit)
                        )
                        
                        modulkombi_df.at[idx, Inhalte_Feld] = inhalt_ai
                    else:
                        modulkombi_df.at[idx, Inhalte_Feld] = "keine Inhalte gefunden"
                        
                    if DEBUG:
                        print(f"Feld 'Bildungsinhalt' erstellt  === Zeile {idx + 1}, jetzt bei {time.time()-start:.2f} sec")
                        log(f"Feld 'Bildungsinhalt' erstellt  === Zeile {idx + 1}")
                
                
                if stop_event.is_set():
                    # print(f"⏹️ Thread {idx + 1} abgebrochen vor Verarbeitung")
                    return
                # --- Zielgruppe ---
                if options.createZielgruppe:
                    pm3 = PromptManager(initial_data=prompts)
                    zielgruppe_zusammen = "\n\n".join(_as_str(x) for x in zielgruppen).strip()
                    if zielgruppe_zusammen is not None and zielgruppe_zusammen.strip() != "" and idx < AnzahlKurse:
                        zielgruppen_clean = clean_zielgruppe_inputs(zielgruppen)
                        ai_ziel = _as_str(SAFE(run_ai_prompt, "ai_zielgruppe", pm3, field_blocks=zielgruppen_clean, model="gpt-4o", max_output_tokens=300, logger=worker.error.emit))
                        final_zielgruppe = f"[KI-Zusammenfassung]\n{ai_ziel}\n\n[Original-Inhalte]\n{zielgruppe_zusammen}"
                        modulkombi_df.at[idx, 'Zielgruppe'] = final_zielgruppe
                    else:
                        modulkombi_df.at[idx, 'Zielgruppe'] = zielgruppe_zusammen
                    if DEBUG:
                        print(f"Feld 'Zielgruppe' erstellt  === Zeile {idx + 1}, jetzt bei {time.time()-start:.2f} sec")
                        log(f"Feld 'Zielgruppe' erstellt  === Zeile {idx + 1}")
                
                
                
                if stop_event.is_set():
                    # print(f"⏹️ Thread {idx + 1} abgebrochen vor Verarbeitung")
                    return
                # --- Voraussetzungen ---
                if options.createVoraussetzungen:
                    pm4 = PromptManager(initial_data=prompts)
                    voraus_zusammen = "\n\n".join(_as_str(x) for x in voraussetzungen).strip()
                    if voraus_zusammen is not None and voraus_zusammen.strip() != "" and idx < AnzahlKurse:
                        ai_voraus = _as_str(SAFE(run_ai_prompt, "ai_voraussetzungen", pm4, field_blocks=voraussetzungen, model="gpt-4o", max_output_tokens=300, logger=worker.error.emit))
                        final_voraus = f"[KI-Zusammenfassung]\n{ai_voraus}\n\n[Original-Inhalte]\n{voraus_zusammen}"
                        modulkombi_df.at[idx, 'Voraussetzungen'] = final_voraus
                    else:
                        modulkombi_df.at[idx, 'Voraussetzungen'] = voraus_zusammen
                    if DEBUG:
                        print(f"Feld 'Voraussetzungen' erstellt  === Zeile {idx + 1}, jetzt bei {time.time()-start:.2f} sec")
                        log(f"Feld 'Voraussetzungen' erstellt  === Zeile {idx + 1}")
    
                
                # Abschlussart
                if options.createAbschlussart:
                    process_abschlussfeld(
                        modulkombi_df, idx, abschlussart, "Abschlussart",
                        "Trägerinterner Abschluss: GFN-Zertifikat",
                        "Trägerinterne Abschlüsse: GFN-Zertifikate",
                        ai_abschlussart,
                        min_length_for_ki=0,
                        max_kurse=AnzahlKurse,
                        worker=worker
                    )
                
                # Abschlussbezeichnung
                if options.createAbschlussbezeichnung:
                    process_abschlussfeld(
                        modulkombi_df, idx, abschlussbezeichnung, "Abschlussbezeichnung",
                        "GFN-Zertifikat", "GFN-Zertifikate",
                        ai_abschlussart,
                        min_length_for_ki=0,
                        max_kurse=AnzahlKurse,
                        worker=worker
                    )
                
                    if DEBUG:
                        print(f"Felder zum Abschluss erstellt  === Zeile {idx + 1}, jetzt bei {time.time()-start:.2f} sec")
                        log(f"Felder zum Abschluss erstellt  === Zeile {idx + 1}")
                
                if stop_event.is_set():
                    # print(f"⏹️ Thread {idx + 1} abgebrochen vor Verarbeitung")
                    return
                # --- Keywords ---
                if options.createKeywords:
                    pm7 = PromptManager(initial_data=prompts)
                    if idx < AnzahlKurse:
                        keywords_table = find_best_keywords(
                            original_title, current_row_title, keywords_df,
                            top_n=30, title_weight=0.5, title_match_weight=0.4,
                            return_keywords_only=True
                        )
                        
                        base_keywords_str = ", ".join(keywords_table)
                        main_keyword = keywords_table[0] if keywords_table else ""
                        
                        raw_str = _as_str(SAFE(run_ai_prompt, "ai_keywords", 
                                               pm7,
                                               titel=original_title,
                                               base_keywords_str=base_keywords_str, 
                                               main_keyword=main_keyword, 
                                               max_output_tokens=500, 
                                               model="gpt-4o", logger=worker.error.emit  ))
                        ai_generated_keywords = [kw.strip() for kw in raw_str.split(",") if kw.strip()]
                        
                        if DEBUG and False:
                            print("60 Keywords: ", ai_generated_keywords)
                        keywords = filter_top_keywords_from_list(titel=selected_title if selected_title else original_title, 
                                                                 base_keywords=split_parts,
                                                                 keywords_a=ai_generated_keywords, df_b=keywords_df, 
                                                                 top_n=30, weight_match=0.6, weight_results=0.3, 
                                                                 match_threshold=0.6, redundancy_threshold=0.85) 
                        modulkombi_df.at[idx, Keyword_Feld] = keywords
                        if DEBUG:
                            print(f"Feld 'Keywords' erstellt  === Zeile {idx + 1}, jetzt bei {time.time()-start:.2f} sec")
                            log(f"Feld 'Keywords' erstellt  === Zeile {idx + 1}")
                            
                    else:
                        modulkombi_df.at[idx, Keyword_Feld] = "keine Keywords gefunden"
                        
                
                
                if stop_event.is_set():
                    # print(f"⏹️ Thread {idx + 1} abgebrochen vor Verarbeitung")
                    return
                # ---------- Systematik ----------
                if options.createSystematik:
                    pm8 = PromptManager(initial_data=prompts)
                    if idx < AnzahlKurse:
                        inhalte_kurz = extract_inhalte_auf_einen_blick(inhalt_ai)
                        print(inhalte_kurz)
                        if not inhalte_kurz:
                            modulkombi_df.at[idx, Systematik_Feld] = "keine Systematik gefunden"
        
                        else:
                            # 1) echte Keywords aus Spalte „Keyword_Feld“ holen
                            kw_raw = keywords if keywords else ai_generated_keywords
                            kw_list = [k.strip() for k in str(kw_raw).split(",") if k.strip()]
                        
                            # 2) Basis = Titel-Teile + echte Keywords (doppelt gewichten)
                            basis_words = split_parts + kw_list[:20]
                            basis_text  = " ".join(basis_words)
                            
                            # print("Basis Text: " + basis_text)
                            
                        
                            # 3) Vorauswahl mit dieser starken Basis
                            short = systematik_vorauswahl(
                                selected_title if selected_title else original_title,
                                basis_text,
                                inhalte_kurz,
                                top=20,
                                gewichtung_titel=5,      
                                max_zusatz_woerter=12,
                                semantic_weight=0.8
                            )
                            
                        
                            # 4) Falls immer noch < 3 Treffer → Basis nochmal erweitern
                            if len(short) < 4:
                                basis_text += " " + " ".join(kw_list[:15])         # nochmal 15 Keywords
                                short = systematik_vorauswahl(
                                    selected_title if selected_title else original_title,
                                    basis_text,
                                    inhalte_kurz,
                                    top=18,
                                    gewichtung_titel=3,
                                    max_zusatz_woerter=15
                                )
                        
                            shorts = "\n".join(str(s).strip() for s in short if s)
                            # print(shorts)
                            best   = SAFE(run_ai_prompt, "llm_best_two_kursnet",
                                          pm8,
                                          titel=original_title,
                                          module=current_row_title,
                                          inhalt=inhalte_kurz,
                                          shortlist=shorts,
                                          model="gpt-4o-mini",
                                          max_output_tokens=200, logger=worker.error.emit)
                        
                            final_best = (f"[KI-Zusammenfassung]\n{best}\n\n"
                                          f"[Liste passender Systematiken]\n{shorts}")
                            modulkombi_df.at[idx, Systematik_Feld] = final_best
                    else:
                        modulkombi_df.at[idx, Systematik_Feld] = "keine Systematik gefunden"
                    
                    if DEBUG:
                        print(f"Feld 'Systematik' erstellt  === Zeile {idx + 1}, jetzt bei {time.time()-start:.2f} sec")
                        log(f"Feld 'Systematik' erstellt  === Zeile {idx + 1}")
                            
    
            else:
                unfound_modules.append((current_row_title, mod_ids))
    
            # ---- Write only the current DF row to Excel (same end result, less I/O) ----
            # Convert NA→None only for this row on write
            if DEBUG:
                print(f"vorm Speichern === Zeile {idx + 1}, jetzt bei {time.time()-start:.2f} sec")
            row_pos = modulkombi_df.index.get_loc(idx)  # 0-based position in DF
            r_idx = row_pos + 2                          # Excel row (header at row 1)
    
            df_row = modulkombi_df.iloc[row_pos]
            with write_lock:  # schützt vor parallelen Excel-Schreibzugriffen
                for c_idx, value in enumerate(df_row.tolist(), start=1):
                    val = None if is_scalar_na(value) else _as_str(value)
                    ws.cell(row=r_idx, column=c_idx, value=val)
    
                # checkpoint alle 10 Zeilen
                if (row_pos + 1) % 10 == 0:
                    print("💾 Speichere Zwischenstand...")
                    success = safe_save(wb, backup_file)
                    if not success:
                        print("⚠️ Speichern übersprungen oder fehlgeschlagen, fahre fort...")
                    print(f"💾 Checkpoint saved at row {r_idx}")
            
            return
                    
    
        except (KeyboardInterrupt):
            print("🛑 Benutzerabbruch erkannt – sichere Fortschritte ...")
            log("🛑 Benutzerabbruch erkannt – sichere Fortschritte ...")
            try:
                success = safe_save(wb, backup_file)
                if not success:
                    msg2 = "⚠️ Speichern übersprungen oder fehlgeschlagen, fahre fort..."
                    worker.warning.emit(msg2)
                    print("⚠️ Speichern übersprungen oder fehlgeschlagen, fahre fort...")
                shutil.move(backup_file, FILENAME)
                worker.progress.emit("✅ Original aktualisiert.")
                print(f"✅ Original aktualisiert: {FILENAME}")
            except Exception as e_inner:
                worker.error.emit(f"Fehler beim Speichern nach Abbruch: {e_inner}\n\n{traceback.format_exc()}")
                print(f"⚠️ Fehler beim Speichern: {e_inner}")
            return
        
        except ModuleNotFoundError as e:          
            print("Fehler bei der Modulzuordnung:", e)
            worker.error.emit(f"Fehler bei der Modulzuordnung: {e}\n\n{traceback.format_exc()}")

    
        except Exception as e:
            tb = traceback.format_exc()
            worker.error.emit(f"❌ Fehler in Zeile {idx + 1}: {e}\n\n{tb}")
            retries += 1
            print(f"❌ Fehler in Zeile {idx + 1}: {e}")
            if retries >= MAX_RETRIES:
                print("❌ Maximale Anzahl an Wiederholungen. Nächste Zeile...")
                # last-ditch: write current row as-is, then exit
                try:
                    success = safe_save(wb, backup_file)
                    if not success:
                        print("⚠️ Speichern übersprungen oder fehlgeschlagen, fahre fort...")
                finally:
                    return
            else:
                print(f"⏳ Waiting {WAIT_SECONDS} seconds before retry...")
                time.sleep(WAIT_SECONDS)




print("Starte GUI...")
print("✅ Alles fertig!")



        
# print("Spalten im DF:", list(modulkombi_df.columns))
# print("Titel-Spalte:", type(TitelNeu), TitelNeu)

# ---------------- Worker (QThread) ----------------
class ProcessWorker(QThread):
    progress = Signal(str)
    finished = Signal()
    error = Signal(str)     # Neues: kritische Fehler (modal)
    warning = Signal(str)   # Neues: weniger kritische Hinweise
    
    ask_module_correction = Signal(object)      # unfound_titles
    ask_keyword = Signal(object)                # (original_title, best_kw_rows, detail_data1)
    ask_title = Signal(object)                  # (original_title, new_titles, detail_data2)
    
    result = None
    result_event = threading.Event()
    
    
    def get_result_from_gui(self, payload):
        """Wird vom GUI aufgerufen"""
        self.result = payload
        self.result_event.set()
    
    def wait_for_gui_result(self):
        if not self.result_event.wait(timeout=600):  # 10 Minuten
            raise RuntimeError("Timeout waiting for GUI dialog result (10min)")
        r = self.result
        self.result = None
        self.result_event.clear()
        return r

    def __init__(self, df, wb, ws, backup_file, options: GenerationOptions):
        super().__init__()
        self.df = df
        self.wb = wb
        self.ws = ws
        self.backup_file = backup_file
        self.result = None
        self.result_event = threading.Event()
        self.stop_event = threading.Event()
        self.options = options

    def run(self):
        self.progress.emit("🚀 Starte Verarbeitung...")
        try:
            with ThreadPoolExecutor(max_workers=1) as executor:
                futures = {
                    executor.submit(
                        process_row, idx, row, self.df, self.ws, self.wb, self.backup_file, self.stop_event, self.options, self
                    ): idx
                    for idx, row in self.df.iterrows()
                }

                try:
                    for future in as_completed(futures):
                        idx = futures[future]
                        try:
                            # blockiert bis Aufgabe fertig/Exception
                            future.result()
                            self.progress.emit(f"✅ Zeile {idx + 1} abgeschlossen")

                        except UserCancelledError:
                            self.progress.emit("🚨 Benutzer hat Abbrechen gedrückt – Fortschritt wird gespeichert ...")
                            success = safe_save(self.wb, self.backup_file)
                            if not success:
                                self.progress.emit("⚠️ Speichern übersprungen oder fehlgeschlagen, fahre fort...")
                            try:
                                shutil.move(self.backup_file, FILENAME)
                            except Exception as e:
                                self.progress.emit(f"⚠️ Fehler beim Verschieben: {e}")
                            self.progress.emit(f"✅ Fortschritt gespeichert bis Reihe {idx + 1}.")
                            break

                        except KeyboardInterrupt:
                            self.progress.emit("🛑 KeyboardInterrupt erkannt – Fortschritt wird gespeichert...")
                            success = safe_save(self.wb, self.backup_file)
                            if not success:
                                self.progress.emit("⚠️ Speichern übersprungen oder fehlgeschlagen, fahre fort...")
                            try:
                                shutil.move(self.backup_file, FILENAME)
                            except Exception as e:
                                self.progress.emit(f"⚠️ Fehler beim Verschieben: {e}")
                            self.progress.emit("✅ Abbruch sauber gespeichert.")
                            break

                        except Exception as e:
                            tb = traceback.format_exc()
                            self.error.emit(f"Unerwarteter Fehler im Hauptprozess: {e}\n\n{tb}")
                            # konkrete Exception-Meldung wie vorher
                            self.progress.emit(f"⚠️ Zeile {idx + 1} fehlgeschlagen: {e}")
                            if DEBUG:
                                traceback.print_exc()

                        # Stop prüfen und ggf. abbrechen:
                        if self.stop_event.is_set():
                            self.progress.emit("🛑 Stop-Event erkannt im Worker; beende weitere Verarbeitung.")
                            break

                finally:
                    executor.shutdown(wait=False)

            # Final Save (einmalig)
            saved = safe_save(self.wb, self.backup_file)
            if saved:
                try:
                    shutil.move(self.backup_file, FILENAME)
                    self.progress.emit(f"✅ Final gespeichert als {FILENAME}")
                except Exception as e:
                    self.progress.emit(f"⚠️ Fehler beim Verschieben: {e}")
            else:
                self.progress.emit("⚠️ Finales Speichern fehlgeschlagen.")

        except Exception as e:
            self.progress.emit(f"❌ Unerwarteter Fehler im Hauptprozess: {e}")
            if DEBUG:
                traceback.print_exc()

        self.progress.emit("🏁 Verarbeitung abgeschlossen.")
        self.finished.emit()

    def stop(self):
        self.stop_event.set()
        self.progress.emit("🛑 Stop-Event gesetzt.")


def run_duplicate_check(df, wb, ws):
    
    if getattr(run_duplicate_check, "_running", False):
        print("⚠️ Dublettenprüfung bereits aktiv – Abbruch.")
        return
    
    run_duplicate_check._running = True

    print("🔍 Starte Dublettenprüfung am Ende...")

    # Titel-Spalte finden
    titel_col = TitelNeu  # <- DEINE VARIABLE

    # Normalisierte Titelspalte
    df["_duplicate_norm"] = df[titel_col].astype(str).str.lower().str.strip()

    dup_mask = df.duplicated("_duplicate_norm", keep=False)
    groups = df[dup_mask].groupby("_duplicate_norm")

    duplicate_groups = [g.drop(columns=["_duplicate_norm"]) for _, g in groups]

    if not duplicate_groups:
        print("✅ Keine Dubletten vorhanden.")
        return

    print(f"📦 {len(duplicate_groups)} Dubletten-Gruppen gefunden – öffne UI...")

    # Mapping erstellen
    colmap = {
        "titel": TitelNeu,
        "original": TitelFeld,
        "modul": Modulnummern,
        "einleitung": EinleitungNeu,
        "inhalte": Inhalte_Feld
    }

    changes = DuplicateResolutionDialog.resolve_duplicates(duplicate_groups, colmap)

    print("🔍 Änderungen:", changes)

    if not changes:
        print("⚠️ Keine Änderungen oder Abbrechen gedrückt – trotzdem Fortschritt gespeichert.")
        return

    # Änderungen in DataFrame UND Excel übernehmen
    for idx, updates in changes.items():
        for col, val in updates.items():
            df.at[idx, col] = val
            r = idx + 2
            c = df.columns.get_loc(col) + 1
            ws.cell(row=r, column=c, value=val)

    print("Dublettenprüfung erfolgreich abgeschlossen.")
    
    safe_save(wb, backup_file) 
    shutil.move(backup_file, FILENAME) 
    print(f"✅ Final gespeichert als {FILENAME}") 
    print(f"\n⏱️ Gesamtzeit: {time.time()-start:.2f} Sekunden")
    run_duplicate_check._running = False



# ---------------- GUI ----------------
class MasterUI(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Master UI – Prozesssteuerung")
        self.setMinimumWidth(520)
        
        # Logs
        log_emitter.log_signal.connect(self._append_log)

        # Layout
        main = QVBoxLayout(self)

        # Buttons
        row = QHBoxLayout()
        self.start_btn = QPushButton("▶️ Starten")
        self.stop_btn = QPushButton("🛑 Stoppen")
        self.stop_btn.setEnabled(False)
        row.addWidget(self.start_btn)
        row.addWidget(self.stop_btn)
        main.addLayout(row)

        self.prompt_manager = PromptManager(initial_data=prompts)

        self.edit_prompts_btn = QPushButton("📝 Prompts bearbeiten")
        main.addWidget(self.edit_prompts_btn)
        self.edit_prompts_btn.clicked.connect(self._open_prompt_editor)


        # Checkboxen
        gb_cb = QGroupBox("Optionen")
        cb_layout = QVBoxLayout()
        self.checkboxes = {}
        opts = [
            "createMasterKeyword", "createTitel", "createEinleitung", "createInhalte",
            "createZielgruppe", "createVoraussetzungen", "createAbschlussart",
            "createAbschlussbezeichnung", "createSystematik", "createKeywords", 
            "createTermine"
        ]
        
        # Anzeige-Namen für UI
        label_map = {
            "createMasterKeyword": "Master-Keyword erzeugen",
            "createTitel": "Titel generieren",
            "createEinleitung": "Einleitung erstellen",
            "createInhalte": "Bildungsinhalt erstellen",
            "createZielgruppe": "Zielgruppe erstellen",
            "createVoraussetzungen": "Voraussetzungen erstellen",
            "createAbschlussart": "Abschlussart erzeugen",
            "createAbschlussbezeichnung": "Abschluss-Bezeichnung erstellen",
            "createSystematik": "Systematik erzeugen",
            "createKeywords": "Keywords generieren",
            "createTermine": "Termine erstellen"
        }
        
        tooltip_map = {
            "createMasterKeyword": "Finde das relevanteste Keyword.",
            "createTitel": "Wähle aus vorgeschlagenen Kurs-Titel.",
            "createEinleitung": "Erzeugt einen kurzen, knackigen Beschreibungstext.",
            "createInhalte": "Erstellt KI-basierten Struktur- und Lerninhaltsblock.",
            "createZielgruppe": "Erzeugt das Feld zu den Zielgruppen automatisch.",
            "createVoraussetzungen": "Erzeugt das Feld zu den Voraussetzungen automatisch.",
            "createAbschlussart": "Legt das Feld Abschlussart fest (z. B. Zertifikat).",
            "createAbschlussbezeichnung": "Generiert die Bezeichnung des Abschlusses.",
            "createSystematik": "Ordnet dem Lehrgebiet eine passende Systematik zu.",
            "createKeywords": "Generiert SEO-relevante Keywords.",
            "createTermine": "Erstellt Termine automatisch."
        }
        
        for opt in opts:
            label = label_map.get(opt, opt)
            cb = QCheckBox(label)
            cb.setChecked(True)
            cb.setToolTip(tooltip_map.get(opt, ""))
            self.checkboxes[opt] = cb
            cb_layout.addWidget(cb)
            
        
        
        gb_cb.setLayout(cb_layout)
        main.addWidget(gb_cb)

        # Status
        self.status = QLabel("Bereit.")
        main.addWidget(self.status)
        
        # Log-Fenster (nicht-editierbar)
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setMinimumHeight(140)
        main.addWidget(self.log)

        # Wire buttons
        self.start_btn.clicked.connect(self.on_start)
        self.stop_btn.clicked.connect(self.on_stop)

        # placeholders for workbook and worker
        self.wb = None
        self.ws = None
        self.df = None
        self.worker = None
    
    def _open_prompt_editor(self):
        dlg = PromptEditorDialog(self.prompt_manager, parent=self)
        dlg.exec()

        
        
    def handle_module_dialog(self, prefill):
        dlg = ModuleEditDialog(prefill, parent=self)
        result = dlg.results()
    
        self.worker.get_result_from_gui(result)
        
    
    def handle_keyword_dialog(self, data):
        result = KeywordSelectionDialog.get_selected_keyword(
            title=data["title"],
            rows=data["rows"],
            keywords_df=data["keywords_df"],
            search_threshold=data.get("search_threshold", 0.3),
            search_top_n=data.get("search_top_n", 25),
            detail_data=data.get("detail_data")
        )
    
        self.worker.get_result_from_gui(result)

        
    def handle_title_dialog(self, data):
        result = TitleSelectionDialog.get_selected_title(
            data["original_title"],
            data["new_titles"],
            data["detail_data"]
        )
    
        self.worker.get_result_from_gui(result)




    def _load_excel(self):
        """Lädt wb/ws und DataFrame; gibt (wb, ws, df) oder wirft Exception."""
        
        # 🛑 Nicht versuchen lokal zu laden, wenn aus Azure
        if FILENAME.startswith("azure:"):
            if "modulkombi" not in OTHER_FILES or "azure_params" not in OTHER_FILES["modulkombi"]:
                raise RuntimeError("Azure-Dateiinfos fehlen.")
            
            params = OTHER_FILES["modulkombi"]["azure_params"]
            sheet = OTHER_FILES["modulkombi"]["sheet"]
        
            # 👉 statt workbook laden: DataFrame aus dem Loader benutzen
            df = modulkombi_df  # wurde ja schon geladen
            return None, None, df
        
        wb = openpyxl.load_workbook(FILENAME)
        if SHEETNAME not in wb.sheetnames:
            raise RuntimeError(f"Sheet '{SHEETNAME}' nicht gefunden in {FILENAME}.")
        ws = wb[SHEETNAME]
        # pandas DataFrame aus Excel (openpyxl engine)
        df = pd.read_excel(FILENAME, sheet_name=SHEETNAME, engine="openpyxl")
        return wb, ws, df

    def on_start(self):
        # Wichtig: stop_event zurücksetzen
        if self.worker:
            self.worker.stop_event.clear()
        # lade Excel vor dem Start — verhindert NoneType.save
        try:
            self.wb, self.ws, self.df = self._load_excel()
        except Exception as e:
            self.status.setText(f"⚠️ Excel-Load fehlgeschlagen: {e}")
            print(f"⚠️ Excel-Load fehlgeschlagen: {e}")
            if DEBUG:
                traceback.print_exc()
            return

        # Backup-file definieren
        backup_file = FILENAME.replace(".xlsx", "_progress.xlsx")

        # Start Worker
        opt_dict = {k: cb.isChecked() for k, cb in self.checkboxes.items()}
        self.options = GenerationOptions(**opt_dict)
        self.worker = ProcessWorker(self.df, self.wb, self.ws, backup_file, options=self.options)
        self.worker.progress.connect(self._on_progress)
        self.worker.warning.connect(self._on_warning)
        self.worker.error.connect(self._on_error)
        self.worker.finished.connect(self._on_finished)
        
        self.worker.ask_module_correction.connect(self.handle_module_dialog)
        self.worker.ask_keyword.connect(self.handle_keyword_dialog)
        self.worker.ask_title.connect(self.handle_title_dialog)

        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self._set_ui_enabled(False)
        
        self.status.setText("⏳ Starte Prozess...")
        self.worker.start()
        
    def _on_error(self, msg: str):
        # 1) Log immer anhängen
        self._append_log("❌ " + msg)
    
        # 2) Modal anzeigen (kritisch)
        QMessageBox.critical(self, "Fehler", msg)
        
    def _on_warning(self, msg: str):
        self._append_log("⚠️ " + msg)
        # optional: kleine nicht-blockierende Hinweis-Box
        # QMessageBox.information(self, "Hinweis", msg)
        
    def _append_log(self, text: str):
        ts = time.strftime("%Y-%m-%d %H:%M:%S")
        self.log.append(f"[{ts}] {text}")

    def on_stop(self):
        if self.worker:
            self.worker.stop()
            self.status.setText("🛑 Stop-Event gesendet")

    def _on_progress(self, msg):
        self.status.setText(msg)
        print(msg)

    def _on_finished(self):
        if getattr(self, "_handled_finish", False):
            return  # ⛔ Schon verarbeitet -> raus
        self._handled_finish = True
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self._set_ui_enabled(True)
        self.status.setText("⏳ Verarbeitung abgeschlossen – führe Dublettenprüfung aus...")
    
        # --- Worker Reset ---
        if self.worker:
            self.worker.stop_event.clear()
            self.worker.result_event.clear()
            self.worker = None
    
        # --- Dublettenprüfung ---
        try:
            run_duplicate_check(self.df, self.wb, self.ws)
        except Exception as e:
            print(f"⚠️ Fehler in der Dublettenprüfung: {e}")
            traceback.print_exc()
            
        
        if self.options.createTermine:
            self.status.setText("⏳ Dublettenprüfung abgeschlossen – erstelle Termine...")
            
            create_termine(self.options, FILENAME, SHEETNAME, "mögliche Zeiträume für ESTHER \n(rot = keine Kombination gefunden, \ngrün = fertig) ")
        else:
            self.status.setText("⏳ Dublettenprüfung abgeschlossen...")    
    
        self.status.setText("✅ Fertig.")
        
    def _set_ui_enabled(self, enabled: bool):        
        self.edit_prompts_btn.setEnabled(enabled)
    
        # Checkboxen^
        for cb in self.checkboxes.values():
            cb.setEnabled(enabled)
    
        # Gruppenboxen deaktivieren
        for gb in self.findChildren(QGroupBox):
            gb.setEnabled(enabled)
    
        # Status bleibt aktiv, Stop wird separat gesteuert
        self.status.setEnabled(True)
    
        # Buttons: Start und Stop werden separat gesetzt
        self.start_btn.setEnabled(enabled)
        # stop_btn bleibt unberührt hier!



# ---------------- Main ----------------
if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    initialize_column_config()
    
    loader = ExcelLoader()
    result = loader.exec()  # blockiert hier, bis Dialog fertig
    
    if result != QDialog.Accepted:
        print("Loader abgebrochen. Programm wird beendet.")
        sys.exit(0)
        
    
    # Prüfe globale Variablen
    missing_final = []
    for varname in ["modulkombi_df", "module_df", "mapping_df", "keywords_df", "SYSTEMATIK"]:
        if globals().get(varname) is None:
            missing_final.append(varname)
    if missing_final:
        print("Nicht alle Dateien geladen:", missing_final)
        sys.exit(1)

    # Backup-Dateiname aus FILENAME
    try:
        backup_file = FILENAME.replace(".xlsx", "_progress.xlsx")
    except Exception:
        backup_file = "backup_progress.xlsx"

    print("Alle Dateien geladen. Weiter mit MasterUI...")
    # 2. Flache Systematik-Liste erzeugen (einmalig)
    SYSTEMATIK_FLAT = list(set([
        word.lower() for systematiken in SYSTEMATIK  # deine volle Systematik-Liste
        for word in re.findall(r'\w+', systematiken) if len(word) > 3
    ]))
    w = MasterUI()
    w.show()
    sys.exit(app.exec())



# ---- tail / wrap-up ---- 
if unfound_modules: 
    print("=== NICHT GEFUNDENE MODULE ===") 
    for kurs, ids in unfound_modules: 
        print(f"Kurs: {kurs}") 
        print(" Nicht gefunden:", ", ".join(_as_str(i) for i in ids)) 
        





        

